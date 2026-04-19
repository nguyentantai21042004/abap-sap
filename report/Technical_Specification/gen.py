import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Technical_Specification.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFCC99")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
CODE_FONT = Font(name="Courier New", size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=8):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
ws["A1"].value = "Technical Specification"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 40
ws.merge_cells("A2:H2")
ws["A2"].value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Function ID", "ZBUG_WS_v5"),
    ("Module", "ABAP Cross-Module"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Program", "Z_BUG_WORKSPACE_MP (Module Pool, Type M)"),
    ("Developer", "DEV-089"),
    ("Reviewed by", "—"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1
widths(ws, [20, 32, 12, 12, 12, 12, 12, 12])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
ws2["A1"].value = "Change History — Technical Specification"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30
r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2,
    r,
    ["1", "1.0", "Initial technical spec for v5.0", "All", "17/04/2026", "DEV-089", ""],
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: Introduction
# ============================================================
ws3 = wb.create_sheet("Introduction")
ws3.merge_cells("A1:F1")
ws3["A1"].value = "Introduction"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

r = 3
for k, v in [
    ("Function ID", "ZBUG_WS_v5"),
    ("Processing Time", "Online (Dynpro-based, immediate response)"),
    ("Processing Type", "Multilingual (EN/VI)"),
    (
        "Introduction",
        "Custom Module Pool program providing centralised bug tracking for SAP development teams. "
        "Built entirely in ABAP 7.70 using modern syntax (inline DATA(), SWITCH, CONV, string templates, @ host variables).",
    ),
    (
        "Supplement",
        "No batch jobs. No RFC calls. No BAdIs. Direct DB operations on 6 custom Z-tables.",
    ),
]:
    cl(ws3, r, 1, k, None, BOLD_FONT, LEFT)
    ws3.merge_cells(f"B{r}:F{r}")
    cl(ws3, r, 2, v, None, BODY_FONT, LEFT)
    ws3.row_dimensions[r].height = 40 if len(v) > 80 else 20
    r += 1
widths(ws3, [20, 60, 12, 12, 12, 12])

# ============================================================
# Sheet 4: Scope
# ============================================================
ws4 = wb.create_sheet("Scope")
ws4.merge_cells("A1:D1")
ws4["A1"].value = "Scope"
ws4["A1"].fill = HDR_FILL
ws4["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
r = section(ws4, r, "In Scope", 4)
hrow(ws4, r, ["No.", "Item", "Description", "Notes"], fill=SUB_FILL)
r += 1
in_scope = [
    (
        "1",
        "Z_BUG_WORKSPACE_MP Module Pool",
        "All 6 includes (TOP, F00, PBO, PAI, F01, F02)",
        "",
    ),
    (
        "2",
        "10 Dynpro Screens",
        "0200, 0210, 0220, 0300, 0310–0360, 0370, 0400, 0410, 0500",
        "",
    ),
    (
        "3",
        "6 Custom DB Tables",
        "ZBUG_TRACKER, ZBUG_USERS, ZBUG_PROJECT, ZBUG_USER_PROJEC, ZBUG_HISTORY, ZBUG_EVIDENCE",
        "",
    ),
    ("4", "ABAP Dictionary Objects", "3 domains, 10+ data elements", ""),
    ("5", "SAP BCS Email Integration", "CL_BCS, CL_DOCUMENT_BCS, CL_SAPUSER_BCS", ""),
    ("6", "SMW0 Web Repository", "3 report templates + 1 project upload template", ""),
    ("7", "SAPScript Long Text", "3 text IDs in ZBUG_NOTE text object", ""),
    (
        "8",
        "GUI Statuses / Title Bars",
        "SE41 statuses + Transaction Code ZBUG_WS (SE93)",
        "",
    ),
]
for rd in in_scope:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, CENTER if c == 1 else LEFT)
    r += 1

r += 1
r = section(ws4, r, "Out of Scope", 4)
hrow(ws4, r, ["No.", "Item", "Reason", "Notes"], fill=SUB_FILL)
r += 1
out_scope = [
    ("1", "Standard SAP Module Configuration", "No SPRO SD/FI/MM — custom Z only", ""),
    ("2", "Fiori / BSP / Web Dynpro UI", "Classic Dynpro (SAP GUI) only", ""),
    (
        "3",
        "Function Group ZBUG_FG",
        "All logic inlined as FORM routines in includes",
        "",
    ),
    (
        "4",
        "ALE / IDoc / RFC Interfaces",
        "Standalone system — no integration interfaces",
        "",
    ),
]
for rd in out_scope:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, CENTER if c == 1 else LEFT)
    r += 1

widths(ws4, [6, 32, 55, 16])

# ============================================================
# Sheet 5: Assumptions
# ============================================================
ws5 = wb.create_sheet("Assumptions")
ws5.merge_cells("A1:D1")
ws5["A1"].value = "Assumptions"
ws5["A1"].fill = HDR_FILL
ws5["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws5["A1"].alignment = CENTER
ws5.row_dimensions[1].height = 30

r = 3
hrow(ws5, r, ["No.", "Assumption", "Impact if Invalid", "Notes"])
r += 1
assumptions = [
    (
        "1",
        "SAP system S40, Client 324, ABAP version 7.70 (SAP_BASIS 770).",
        "Syntax errors — modern ABAP (inline DATA, SWITCH) unavailable",
        "",
    ),
    (
        "2",
        "Development account DEV-089 has SE11, SE38, SE51, SE41, SE93, SM30 access.",
        "Cannot create/activate ABAP objects",
        "",
    ),
    (
        "3",
        "Users DEV-089, DEV-061, DEV-118 are pre-registered in ZBUG_USERS with roles M, D, T respectively.",
        "Login fails at startup — User not registered error (MSG 010)",
        "",
    ),
    (
        "4",
        "Test projects and mock user data exist in system before UAT testing.",
        "UAT test cases cannot be executed",
        "",
    ),
    (
        "5",
        "SAP email system (SAPConnect) is configured — BCS API can send external emails.",
        "Email notification feature silently fails",
        "",
    ),
    (
        "6",
        "SMW0 web repository is accessible for template upload.",
        "Template download (FORM download_template) returns error",
        "",
    ),
    (
        "7",
        "Status migration script (6->V) will be run exactly once, after v5.0 deployment.",
        "Old data with STATUS=6 (Resolved in v4.x) will be misread as FinalTesting in v5.0",
        "",
    ),
]
for rd in assumptions:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws5, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws5, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws5, r, 3, rd[2], f, BODY_FONT, LEFT)
    cl(ws5, r, 4, rd[3], f, BODY_FONT, LEFT)
    ws5.row_dimensions[r].height = 35
    r += 1
widths(ws5, [6, 55, 45, 16])

# ============================================================
# Sheet 6: Functional Requirements
# ============================================================
ws6 = wb.create_sheet("Functional Requirements")
ws6.merge_cells("A1:E1")
ws6["A1"].value = "Functional Requirements Mapping"
ws6["A1"].fill = HDR_FILL
ws6["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws6["A1"].alignment = CENTER
ws6.row_dimensions[1].height = 30

r = 3
hrow(
    ws6,
    r,
    ["No.", "Functional Requirement", "Technical Implementation", "Include", "Notes"],
)
r += 1
reqs = [
    (
        "1",
        "3-role system (M/D/T)",
        "Read from ZBUG_USERS.ROLE, stored in gv_role",
        "Z_BUG_WS_TOP",
        "",
    ),
    (
        "2",
        "10-state bug lifecycle",
        "Status constants gc_st_*, transition matrix in FORM validate_transition",
        "Z_BUG_WS_F01",
        "",
    ),
    ("3", "Auto-assign Developer", "FORM auto_assign_developer", "Z_BUG_WS_F01", ""),
    ("4", "Auto-assign Tester", "FORM auto_assign_tester", "Z_BUG_WS_F01", ""),
    (
        "5",
        "Status via popup only",
        "Screen group STS locks STATUS field; fcode STATUS_CHG opens screen 0370",
        "Z_BUG_WS_PBO/PAI",
        "",
    ),
    (
        "6",
        "Evidence upload",
        "FORM upload_evidence — writes RAWSTRING to ZBUG_EVIDENCE",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "7",
        "Evidence download",
        "FORM download_evidence — reads RAWSTRING, writes to frontend",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "8",
        "Long text (3 types)",
        "CL_GUI_TEXTEDIT in CC_DESC / CC_DEVNOTE / CC_TSTRNOTE",
        "Z_BUG_WS_F00/F01",
        "",
    ),
    (
        "9",
        "Email notification",
        "CL_BCS, CL_DOCUMENT_BCS, CL_SAPUSER_BCS",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "10",
        "ALV Grid (4 lists)",
        "CL_GUI_ALV_GRID — Bug List, Project List, Evidence List, History List",
        "Z_BUG_WS_F00",
        "",
    ),
    (
        "11",
        "Dashboard Header",
        "Aggregation query on ZBUG_TRACKER, display in text fields",
        "Z_BUG_WS_PBO",
        "",
    ),
    (
        "12",
        "Bug Search (popup)",
        "Screen 0210 input -> FORM execute_bug_search -> Screen 0220 results",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "13",
        "Project Search",
        "Screen 0410 input -> FORM execute_project_search -> Screen 0400",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "14",
        "Template download",
        "FORM download_template reading from SMW0 (WWWDATA)",
        "Z_BUG_WS_F02",
        "",
    ),
    (
        "15",
        "History logging",
        "FORM log_history writes to ZBUG_HISTORY on every change",
        "Z_BUG_WS_F01",
        "",
    ),
    (
        "16",
        "Unsaved changes detection",
        "Global flag gv_data_changed, checked before BACK/EXIT",
        "Z_BUG_WS_PAI",
        "",
    ),
    ("17", "F4 calendar", "FORM f4_date using POPUP_GET_VALUES", "Z_BUG_WS_F02", ""),
    (
        "18",
        "Soft delete",
        "IS_DEL = 'X', never physically delete rows",
        "Z_BUG_WS_F01",
        "",
    ),
]
for rd in reqs:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws6, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws6, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws6, r, 3, rd[2], f, BODY_FONT, LEFT)
    cl(ws6, r, 4, rd[3], f, BODY_FONT, CENTER)
    cl(ws6, r, 5, rd[4], f, BODY_FONT, LEFT)
    ws6.row_dimensions[r].height = 25
    r += 1
widths(ws6, [6, 30, 55, 18, 14])
ws6.freeze_panes = "A4"

# ============================================================
# Sheet 7: Technical Design
# ============================================================
ws7 = wb.create_sheet("Technical Design")
ws7.merge_cells("A1:F1")
ws7["A1"].value = "Technical Design"
ws7["A1"].fill = HDR_FILL
ws7["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws7["A1"].alignment = CENTER
ws7.row_dimensions[1].height = 30

r = 3
r = section(ws7, r, "7.1 Business Process Architecture (Layer Overview)", 6)

arch_lines = [
    "Presentation Layer (Dynpro Screens)",
    "  0410 — Project Search",
    "  0400 — Project List (ALV)",
    "  0200 — Bug List (ALV + Dashboard)",
    "  0210/0220 — Bug Search popup + results",
    "  0300 — Bug Detail (Tab Strip: subscreens 0310-0360)",
    "  0370 — Status Transition popup (Modal Dialog)",
    "  0500 — Project Detail",
    "",
    "Application Layer (ABAP Includes)",
    "  Z_BUG_WS_TOP  — Types, global variables, constants",
    "  Z_BUG_WS_F00  — LCL_EVENT_HANDLER (ALV events), field catalogs",
    "  Z_BUG_WS_PBO  — Module MODULE_*_PBO (screen population logic)",
    "  Z_BUG_WS_PAI  — Module MODULE_*_PAI (user command processing)",
    "  Z_BUG_WS_F01  — FORM routines (CRUD, lifecycle, auto-assign, email, history)",
    "  Z_BUG_WS_F02  — FORM routines (F4, long text editor, popup, template DL)",
    "",
    "Data Layer (Custom Z-Tables in ZBUGTRACK package)",
    "  ZBUG_TRACKER      — 29 fields — core bug data",
    "  ZBUG_USERS        — 12 fields — user registry + roles",
    "  ZBUG_PROJECT      — 16 fields — project data",
    "  ZBUG_USER_PROJEC  — 10 fields — user-project assignment",
    "  ZBUG_HISTORY      — 10 fields — audit trail",
    "  ZBUG_EVIDENCE     — 11 fields — binary file storage",
]
for line in arch_lines:
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws7.cell(row=r, column=1, value=line)
    c.font = CODE_FONT
    c.alignment = LEFT
    c.border = BORDER
    ws7.row_dimensions[r].height = 14
    r += 1

r += 1
r = section(ws7, r, "7.2 WBS & Timeline", 6)
hrow(ws7, r, ["Phase", "Content", "Status", "", "", ""], fill=SUB_FILL)
ws7.merge_cells(f"D{r}:F{r}")
r += 1
for ph, content, status in [
    ("A", "Database Hardening (SE11 objects)", "Done"),
    ("B", "Business Logic update (includes)", "Unconfirmed"),
    ("C+D", "Module Pool UI + Advanced Features (v4.2)", "Done"),
    ("E", "Testing (UAT Round 1 — 11 bugs found)", "Done"),
    ("F", "v5.0 Enhancement (CODE complete, deployment pending)", "In Progress"),
]:
    f = GREEN if status == "Done" else (YELLOW if status == "In Progress" else ORANGE)
    cl(ws7, r, 1, ph, f, BOLD_FONT, CENTER)
    cl(ws7, r, 2, content, f, BODY_FONT, LEFT)
    ws7.merge_cells(f"C{r}:F{r}")
    cl(ws7, r, 3, status, f, BOLD_FONT, CENTER)
    ws7.row_dimensions[r].height = 22
    r += 1

r += 1
r = section(ws7, r, "7.3 Data Dictionary Objects", 6)
hrow(
    ws7,
    r,
    ["Object Type", "Name", "Type / Length", "Description", "", ""],
    fill=SUB_FILL,
)
ws7.merge_cells(f"E{r}:F{r}")
r += 1
dd_objects = [
    ("Domain", "zde_bug_status", "CHAR 20", "Bug status values: 1/2/3/4/5/6/7/W/R/V"),
    ("Domain", "zde_sap_module", "CHAR 20", "SAP module: FI/MM/SD/ABAP/BASIS"),
    ("Domain", "zde_bug_role", "CHAR 1", "User role: M/D/T"),
    ("Data Element", "ZDE_BUG_ID", "CHAR 10", "Bug identifier"),
    ("Data Element", "ZDE_PROJECT_ID", "CHAR 20", "Project identifier"),
    ("Data Element", "ZDE_USERNAME", "CHAR 12", "SAP logon name"),
    ("Data Element", "ZDE_BUG_TITLE", "CHAR 100", "Bug title"),
    ("Data Element", "ZDE_BUG_DESC", "STRING", "Bug description"),
    ("Data Element", "ZDE_REASONS", "STRING", "Root cause text"),
    ("Data Element", "ZDE_PRJ_NAME", "CHAR 100", "Project name"),
    ("Data Element", "ZDE_IS_DEL", "CHAR 1", "Soft delete flag"),
    ("Package", "ZBUGTRACK", "—", "All Z objects belong here"),
    ("Table", "ZBUG_TRACKER", "29 fields", "Bug records"),
    ("Table", "ZBUG_USERS", "12 fields", "User registry"),
    ("Table", "ZBUG_PROJECT", "16 fields", "Projects"),
    ("Table", "ZBUG_USER_PROJEC", "10 fields", "User-Project M:N"),
    ("Table", "ZBUG_HISTORY", "10 fields", "History log"),
    ("Table", "ZBUG_EVIDENCE", "11 fields", "Binary evidence"),
]
for rd in dd_objects:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws7, r, 1, rd[0], f, BODY_FONT, LEFT)
    cl(ws7, r, 2, rd[1], f, CODE_FONT, LEFT)
    cl(ws7, r, 3, rd[2], f, BODY_FONT, CENTER)
    ws7.merge_cells(f"D{r}:F{r}")
    cl(ws7, r, 4, rd[3], f, BODY_FONT, LEFT)
    ws7.row_dimensions[r].height = 20
    r += 1

widths(ws7, [16, 22, 16, 40, 12, 12])
ws7.freeze_panes = "A4"

# ============================================================
# Sheet 8: Development Standards
# ============================================================
ws8 = wb.create_sheet("Development Standards")
ws8.merge_cells("A1:E1")
ws8["A1"].value = "Development Standards"
ws8["A1"].fill = HDR_FILL
ws8["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws8["A1"].alignment = CENTER
ws8.row_dimensions[1].height = 30

r = 3
r = section(ws8, r, "8.1 Naming Conventions", 5)
hrow(ws8, r, ["Object Type", "Prefix", "Example", "", ""], fill=SUB_FILL)
ws8.merge_cells(f"D{r}:E{r}")
r += 1
naming = [
    ("Tables", "ZBUG_", "ZBUG_TRACKER"),
    ("Programs", "Z_BUG_", "Z_BUG_WORKSPACE_MP"),
    ("Includes", "Z_BUG_WS_", "Z_BUG_WS_F01"),
    ("Domains", "zde_", "zde_bug_status"),
    ("Data Elements", "ZDE_", "ZDE_BUG_ID"),
    ("Constants (global)", "gc_", "gc_st_new, gc_role_manager"),
    ("Global variables", "gv_", "gv_role, gv_mode"),
    ("Global structures", "gs_", "gs_bug, gs_project"),
    ("Global internal tables", "gt_", "gt_bugs, gt_projects"),
    ("Local variables", "lv_", "lv_count"),
    ("Local structures", "ls_", "ls_user"),
    ("Local internal tables", "lt_", "lt_evidence"),
    ("Form routines", "snake_case", "auto_assign_developer"),
    ("Screen containers", "CC_", "CC_BUG_LIST, CC_DESC"),
]
for rd in naming:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws8, r, 1, rd[0], f, BODY_FONT, LEFT)
    cl(ws8, r, 2, rd[1], f, CODE_FONT, CENTER)
    ws8.merge_cells(f"C{r}:E{r}")
    cl(ws8, r, 3, rd[2], f, CODE_FONT, LEFT)
    ws8.row_dimensions[r].height = 20
    r += 1

r += 1
r = section(ws8, r, "8.2 ABAP Coding Standards", 5)
hrow(ws8, r, ["Rule", "Standard", "", "", ""], fill=SUB_FILL)
ws8.merge_cells(f"C{r}:E{r}")
r += 1
coding_standards = [
    (
        "ABAP Version",
        "ABAP 7.70 — use inline DATA(), FIELD-SYMBOL, SWITCH, CONV, VALUE, REDUCE",
    ),
    ("String templates", "|Bug { lv_id } saved| — no CONCATENATE"),
    ("Host variables", "@lv_var in all OPEN SQL"),
    ("Obsolete syntax", "No MOVE, COMPUTE, WRITE TO — use modern assignment operators"),
    ("SELECT", "No SELECT * — always specify field list"),
    ("Error handling", "TRY ... CATCH cx_root INTO lx_error"),
    ("Soft delete", "Set IS_DEL = 'X', never use DELETE on Z-tables"),
    ("Auto IDs", "SELECT MAX(bug_id) + 1 pattern (no number range object)"),
]
for rd in coding_standards:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws8, r, 1, rd[0], f, BOLD_FONT, LEFT)
    ws8.merge_cells(f"B{r}:E{r}")
    cl(ws8, r, 2, rd[1], f, BODY_FONT, LEFT)
    ws8.row_dimensions[r].height = 22
    r += 1

r += 1
r = section(ws8, r, "8.3 Database Access Pattern (ABAP Code Samples)", 5)

code_blocks = [
    (
        "Read single record:",
        "SELECT SINGLE bug_id, title, status, dev_id\n"
        "  FROM zbug_tracker\n"
        "  INTO @DATA(ls_bug)\n"
        "  WHERE bug_id = @lv_bug_id\n"
        "    AND is_del <> 'X'.",
    ),
    (
        "Read list with join:",
        "SELECT bt~bug_id, bt~title, bt~status, bu~full_name AS dev_name\n"
        "  FROM zbug_tracker AS bt\n"
        "  LEFT JOIN zbug_users AS bu ON bu~user_id = bt~dev_id\n"
        "  INTO TABLE @DATA(lt_bugs)\n"
        "  WHERE bt~project_id = @lv_project_id\n"
        "    AND bt~is_del <> 'X'.",
    ),
]
for label, code in code_blocks:
    cl(ws8, r, 1, label, SUB_FILL, SUB_FONT, LEFT)
    ws8.merge_cells(f"B{r}:E{r}")
    cl(ws8, r, 2, "", SUB_FILL, SUB_FONT, LEFT)
    ws8.row_dimensions[r].height = 18
    r += 1
    ws8.merge_cells(f"A{r}:E{r}")
    c = ws8.cell(row=r, column=1, value=code)
    c.font = CODE_FONT
    c.alignment = LEFT
    c.border = BORDER
    c.fill = ALT_FILL
    ws8.row_dimensions[r].height = 80
    r += 1

widths(ws8, [24, 24, 20, 20, 20])

# ============================================================
# Sheet 9: Screen Layout
# ============================================================
ws9 = wb.create_sheet("Screen Layout")
ws9.merge_cells("A1:G1")
ws9["A1"].value = "Screen Layout — Key Technical Screens"
ws9["A1"].fill = HDR_FILL
ws9["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws9["A1"].alignment = CENTER
ws9.row_dimensions[1].height = 30

r = 3
r = section(ws9, r, "Screen 0410 — Project Search (Normal, 800x600)", 7)
hrow(ws9, r, ["No.", "Element", "Type", "Name", "Position", "Notes", ""])
ws9.merge_cells(f"G{r}:G{r}")
r += 1
for rd in [
    ("1", "Label", "Text", '"Project ID:"', "Row 3, Col 2", "Static text"),
    (
        "2",
        "Input",
        "OKCODE-related",
        "GV_SEARCH_PRJ_ID",
        "Row 3, Col 15",
        "Search filter",
    ),
    ("3", "Label", "Text", '"Manager:"', "Row 5, Col 2", "Static text"),
    ("4", "Input", "Field", "GV_SEARCH_MGR", "Row 5, Col 15", "Search filter"),
    ("5", "Label", "Text", '"Status:"', "Row 7, Col 2", "Static text"),
    ("6", "Input", "Field", "GV_SEARCH_PRJ_STATUS", "Row 7, Col 15", "Search filter"),
    ("7", "Status Bar", "—", "PBAR", "Bottom", "Standard status bar"),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws9, r, c, v, f, BODY_FONT, CENTER if c == 1 else LEFT)
    ws9.row_dimensions[r].height = 22
    r += 1

r += 1
r = section(ws9, r, "Screen 0370 — Status Transition Popup (Modal Dialog, 600x400)", 7)
hrow(ws9, r, ["No.", "Element", "Type", "Name", "Mode", "Notes", ""])
ws9.merge_cells(f"G{r}:G{r}")
r += 1
for rd in [
    ("1", "Text", "READ", "GV_POPUP_BUG_ID", "Read-only", "Bug ID — always locked"),
    ("2", "Text", "READ", "GV_POPUP_TITLE", "Read-only", "Title — always locked"),
    ("3", "Text", "READ", "GV_POPUP_REPORTER", "Read-only", "Reporter — always locked"),
    (
        "4",
        "Text",
        "READ",
        "GV_POPUP_CUR_STATUS",
        "Read-only",
        "Current status — always locked",
    ),
    (
        "5",
        "Dropdown",
        "INPUT",
        "GV_POPUP_NEW_STATUS",
        "Input",
        "Only allowed transitions shown",
    ),
    (
        "6",
        "Input",
        "Field",
        "GV_POPUP_DEV_ID",
        "Conditional",
        "Enabled when -> Assigned(2)",
    ),
    (
        "7",
        "Input",
        "Field",
        "GV_POPUP_TESTER_ID",
        "Conditional",
        "Enabled when -> FinalTesting(6)",
    ),
    (
        "8",
        "Custom Ctrl",
        "Container",
        "CC_TRANS_NOTE",
        "Always",
        "Long text editor for transition note",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws9, r, c, v, f, BODY_FONT, CENTER if c == 1 else LEFT)
    ws9.row_dimensions[r].height = 22
    r += 1

widths(ws9, [6, 14, 16, 26, 14, 36, 12])
ws9.freeze_panes = "A4"

# ============================================================
# Sheet 10: Screen Definition
# ============================================================
ws10 = wb.create_sheet("Screen Definition")
ws10.merge_cells("A1:G1")
ws10["A1"].value = "Screen Field Definition"
ws10["A1"].fill = HDR_FILL
ws10["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws10["A1"].alignment = CENTER
ws10.row_dimensions[1].height = 30

r = 3
r = section(ws10, r, "Screen 0410 — Project Search Fields", 7)
hrow(
    ws10,
    r,
    ["No.", "Field Name", "Field Label", "Type", "Mandatory", "Screen Group", "Notes"],
)
r += 1
for rd in [
    (
        "1",
        "GV_SEARCH_PRJ_ID",
        "Project ID",
        "CHAR 20",
        "N",
        "—",
        "Search filter — optional",
    ),
    ("2", "GV_SEARCH_MGR", "Manager", "CHAR 12", "N", "—", "Search filter — optional"),
    (
        "3",
        "GV_SEARCH_PRJ_STATUS",
        "Status",
        "CHAR 20",
        "N",
        "—",
        "Search filter — optional",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws10, r, c, v, f, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws10, r, "Screen 0370 — Status Transition Popup Fields", 7)
hrow(
    ws10, r, ["No.", "Field Name", "Label", "Type", "Read/Input", "Condition", "Notes"]
)
r += 1
for rd in [
    ("1", "GV_POPUP_BUG_ID", "Bug ID", "CHAR 10", "Read-only", "Always", "Locked"),
    ("2", "GV_POPUP_TITLE", "Title", "CHAR 100", "Read-only", "Always", "Locked"),
    (
        "3",
        "GV_POPUP_REPORTER",
        "Reporter",
        "CHAR 12",
        "Read-only",
        "Always",
        "= TESTER_ID",
    ),
    (
        "4",
        "GV_POPUP_CUR_STATUS",
        "Current Status",
        "CHAR 20 (domain)",
        "Read-only",
        "Always",
        "Locked",
    ),
    (
        "5",
        "GV_POPUP_NEW_STATUS",
        "New Status",
        "CHAR 20 (domain)",
        "Input (ddlb)",
        "Always",
        "Only valid transitions",
    ),
    (
        "6",
        "GV_POPUP_DEV_ID",
        "Developer ID",
        "CHAR 12",
        "Input",
        "Only when -> Assigned(2)",
        "Mandatory if ->2",
    ),
    (
        "7",
        "GV_POPUP_TESTER_ID",
        "Final Tester",
        "CHAR 12",
        "Input",
        "Only when -> FinalTesting(6)",
        "Mandatory if ->6",
    ),
    (
        "8",
        "CC_TRANS_NOTE",
        "Transition Note",
        "Custom Control",
        "Input",
        "Mandatory for R, V",
        "Long text editor",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws10, r, c, v, f, BODY_FONT, LEFT)
    r += 1

widths(ws10, [6, 24, 18, 18, 14, 28, 20])
ws10.freeze_panes = "A4"

# ============================================================
# Sheet 11: Message Definition
# ============================================================
ws11 = wb.create_sheet("Message Definition")
ws11.merge_cells("A1:F1")
ws11["A1"].value = "Message Definition — Message Class ZBUG_MSG"
ws11["A1"].fill = HDR_FILL
ws11["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws11["A1"].alignment = CENTER
ws11.row_dimensions[1].height = 30

r = 3
hrow(ws11, r, ["Message ID", "Type", "Text (EN)", "Trigger", "Variable", "Notes"])
r += 1
msgs = [
    ("001", "S", "Bug &1 saved successfully", "Bug save OK", "&1 = BUG_ID", ""),
    ("002", "E", "Title is required", "Title field empty on save", "—", ""),
    (
        "003",
        "E",
        "Evidence required before setting Fixed",
        "Moving to status 5 without evidence",
        "—",
        "",
    ),
    (
        "004",
        "E",
        "Transition note required for this status change",
        "Moving to R/V without TRANS_NOTE",
        "—",
        "",
    ),
    (
        "005",
        "S",
        "Status changed: &1 -> &2",
        "Status transition OK",
        "&1=old, &2=new status",
        "",
    ),
    (
        "006",
        "W",
        "No available Developer for module &1 — Bug set to Waiting",
        "Auto-assign fails",
        "&1 = SAP_MODULE",
        "",
    ),
    (
        "007",
        "S",
        "Developer &1 auto-assigned",
        "Auto-assign success",
        "&1 = DEV_ID",
        "",
    ),
    (
        "008",
        "E",
        "Cannot mark project Done: &1 open bug(s) exist",
        "Project closure validation",
        "&1 = count",
        "",
    ),
    ("009", "S", "Project &1 saved", "Project save OK", "&1 = PROJECT_ID", ""),
    (
        "010",
        "E",
        "User &1 not registered in ZBUG_USERS",
        "Invalid user at startup",
        "&1 = user name",
        "",
    ),
    (
        "011",
        "E",
        "Access denied — role &1 cannot perform this action",
        "Permission check failed",
        "&1 = role",
        "",
    ),
    ("012", "S", "Email notification sent", "BCS email OK", "—", ""),
    (
        "013",
        "E",
        "Cannot delete: project has &1 active bug(s)",
        "Project delete validation",
        "&1 = count",
        "",
    ),
]
for rd in msgs:
    f = ALT_FILL if r % 2 == 0 else None
    type_fill = GREEN if rd[1] == "S" else (YELLOW if rd[1] == "W" else f)
    for c, v in enumerate(rd, 1):
        fill = type_fill if c == 2 else f
        cl(ws11, r, c, v, fill, BODY_FONT, CENTER if c in (1, 2) else LEFT)
    r += 1
widths(ws11, [12, 8, 45, 35, 18, 16])
ws11.freeze_panes = "A4"

# ============================================================
# Sheet 12: Technical Implementation
# ============================================================
ws12 = wb.create_sheet("Technical Implementation")
ws12.merge_cells("A1:E1")
ws12["A1"].value = "Technical Implementation Notes"
ws12["A1"].fill = HDR_FILL
ws12["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws12["A1"].alignment = CENTER
ws12.row_dimensions[1].height = 30


def code_block(ws, start_r, label, code, ncols=5):
    ws.merge_cells(start_row=start_r, start_column=1, end_row=start_r, end_column=ncols)
    cl(ws, start_r, 1, label, SUB_FILL, SUB_FONT, LEFT)
    ws.row_dimensions[start_r].height = 20
    lines = code.split("\n")
    for line in lines:
        start_r += 1
        ws.merge_cells(
            start_row=start_r, start_column=1, end_row=start_r, end_column=ncols
        )
        c = ws.cell(row=start_r, column=1, value=line)
        c.font = CODE_FONT
        c.alignment = LEFT
        c.border = BORDER
        c.fill = ALT_FILL
        ws.row_dimensions[start_r].height = 14
    return start_r + 2


r = 3
r = code_block(
    ws12,
    r,
    "11.1 Include Dependency Order",
    """PROGRAM z_bug_workspace_mp.
INCLUDE z_bug_ws_top.   " 1. Global data — MUST be first
INCLUDE z_bug_ws_f00.   " 2. LCL_EVENT_HANDLER class — MUST be before PBO/PAI
INCLUDE z_bug_ws_pbo.   " 3. PBO modules
INCLUDE z_bug_ws_pai.   " 4. PAI modules
INCLUDE z_bug_ws_f01.   " 5. Business logic FORMs
INCLUDE z_bug_ws_f02.   " 6. Helper FORMs""",
)

r = code_block(
    ws12,
    r,
    "11.2 ALV Grid Initialisation Pattern",
    """" Create container + ALV in PBO
IF go_alv_bugs IS INITIAL.
  CREATE OBJECT go_container_bugs
    EXPORTING container_name = 'CC_BUG_LIST'.
  CREATE OBJECT go_alv_bugs
    EXPORTING i_parent = go_container_bugs.
  SET HANDLER go_handler->on_double_click FOR go_alv_bugs.
ENDIF.
CALL METHOD go_alv_bugs->set_table_for_first_display
  EXPORTING i_structure_name = 'ZS_BUG_DISPLAY'
  CHANGING  it_outtab = gt_bugs.""",
)

r = code_block(
    ws12,
    r,
    "11.3 Status Transition Validation (FORM validate_transition)",
    """FORM validate_transition
  USING iv_current TYPE zde_bug_status
        iv_new     TYPE zde_bug_status
        iv_role    TYPE zde_bug_role
  CHANGING cv_valid TYPE abap_bool.

  cv_valid = abap_false.
  CASE iv_current.
    WHEN gc_st_new.          " 1 -> 2 or W (Manager only)
      IF iv_role = gc_role_manager AND iv_new CA '2W'. cv_valid = abap_true. ENDIF.
    WHEN gc_st_assigned.     " 2 -> 3 or R
      IF iv_new = gc_st_inprogress OR iv_new = gc_st_rejected. cv_valid = abap_true. ENDIF.
    WHEN gc_st_inprogress.   " 3 -> 5, 4, R
      IF iv_new CA '54R'. cv_valid = abap_true. ENDIF.
    WHEN gc_st_pending.      " 4 -> 2 (Manager only)
      IF iv_role = gc_role_manager AND iv_new = gc_st_assigned. cv_valid = abap_true. ENDIF.
    WHEN gc_st_finaltesting. " 6 -> V or 3
      IF iv_new = gc_st_resolved OR iv_new = gc_st_inprogress. cv_valid = abap_true. ENDIF.
    WHEN gc_st_waiting.      " W -> 2 or 6 (Manager only)
      IF iv_role = gc_role_manager AND iv_new CA '26'. cv_valid = abap_true. ENDIF.
  ENDCASE.
ENDFORM.""",
)

r = code_block(
    ws12,
    r,
    "11.4 Evidence Upload to ZBUG_EVIDENCE",
    """DATA: lv_file  TYPE string,
      lt_data  TYPE STANDARD TABLE OF raw255,
      ls_evd   TYPE zbug_evidence.

CALL METHOD cl_gui_frontend_services=>file_open_dialog
  CHANGING file_table = DATA(lt_files).

lv_file = lt_files[ 1 ].
cl_gui_frontend_services=>gui_upload(
  filename = lv_file
  filetype = 'BIN'
  CHANGING data_tab = lt_data ).

IMPORT DATA FROM INTERNAL TABLE lt_data TO ls_evd-content.
ls_evd-bug_id       = gs_bug-bug_id.
ls_evd-file_name    = lv_file.
ls_evd-uploaded_by  = sy-uname.
ls_evd-upload_date  = sy-datum.
INSERT zbug_evidence FROM ls_evd.""",
)

r = code_block(
    ws12,
    r,
    "11.5 Email Notification via BCS API",
    """DATA: lo_send_req TYPE REF TO cl_bcs,
      lo_doc      TYPE REF TO cl_document_bcs,
      lo_addr     TYPE REF TO cl_sapuser_bcs.

lo_send_req = cl_bcs=>create_instance( ).
lo_doc = cl_document_bcs=>create_document(
           i_type    = 'RAW'
           i_subject = |Bug { gs_bug-bug_id } status: { lv_new_status }|
           i_text    = lt_body ).
lo_send_req->set_document( lo_doc ).
lo_addr = cl_sapuser_bcs=>create_instance( i_user = gs_bug-dev_id ).
lo_send_req->add_recipient( i_recipient = lo_addr ).
lo_send_req->send( ).
COMMIT WORK.""",
)

widths(ws12, [28, 20, 20, 20, 20])

wb.save(OUT)
print(f"Created: {OUT}")
