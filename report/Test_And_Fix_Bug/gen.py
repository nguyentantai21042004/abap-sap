import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Test_And_Fix_Bug.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFCC99")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 30


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


def sev_fill(sev):
    if sev == "CRITICAL":
        return RED
    if sev == "High":
        return ORANGE
    if sev == "Medium":
        return YELLOW
    return ALT_FILL


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Fix and bugs (Main bug tracker)
# ============================================================
ws = wb.active
ws.title = "Fix and bugs"

ws.merge_cells("A1:J1")
ws["A1"].value = "Test And Fix Bug — UAT Round 1 Bug Tracker"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:J2")
ws[
    "A2"
].value = "Z_BUG_WORKSPACE_MP v5.0 | SAP S40 Client 324 | Phase: SIT/UAT Round 1 | Date: 17/04/2026"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
ws["A2"].alignment = CENTER

r = 4
r = section(ws, r, "Document Information", 10)
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Test Phase", "SIT + UAT Round 1"),
    ("Total Bugs Logged", "11"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
    ("Source", "UAT Round 1 testing on S40 Client 324"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:J{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "Severity Classification", 10)
hrow(ws, r, ["Severity", "Description", "SLA", "", "", "", "", "", "", ""])
ws.merge_cells(f"D{r}:J{r}")
r += 1
for sev, desc, sla in [
    (
        "CRITICAL",
        "Short dump / system crash — blocks all testing",
        "Fix before re-test",
    ),
    (
        "High",
        "Functional defect — incorrect behavior, data loss risk",
        "Fix before UAT Round 2",
    ),
    ("Medium", "UX/design issue — workaround exists", "Fix before go-live"),
    ("Low", "Minor cosmetic issue", "Post go-live"),
]:
    f = sev_fill(sev)
    cl(ws, r, 1, sev, f, BOLD_FONT, CENTER)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, desc, f, BODY_FONT, LEFT)
    ws.merge_cells(f"I{r}:J{r}")
    cl(ws, r, 9, sla, f, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "Bug Tracker — UAT Round 1 (13/04/2026)", 10)
hrow(
    ws,
    r,
    [
        "No.",
        "Bug Title",
        "Reproduction Steps",
        "Expected Result",
        "Actual Result",
        "Root Cause",
        "Fix Applied",
        "Evidence",
        "Severity",
        "Status",
    ],
)
ws.row_dimensions[r].height = 35
r += 1

bugs = [
    (
        "1",
        "Short dump on tab Description/Dev Note/Tester Note",
        "1. Open BUG0000023 in Change mode (Screen 0300)\n2. Click tab Description or Dev Note or Tester Note\n3. ABAP short dump immediately",
        "Tab opens, text editor displays existing content",
        "CALL_FUNCTION_CONFLICT_TYPE short dump. Program terminates.",
        "Custom Controls CC_DESC/CC_DEVNOTE/CC_TSTRNOTE not created on screens 0320/0330/0340 in SE51; or STRING field placed on screen layout causing deep type crash",
        "CODE_PBO: Add TRY-CATCH around CREATE OBJECT go_cont_desc. CODE_F02: Explicit type cast lv_tdname TYPE tdobname before READ_TEXT. SE51: Verify CC_DESC on 0320, CC_DEVNOTE on 0330, CC_TSTRNOTE on 0340.",
        "runtime-error-read-text-type-conflict-115802.png",
        "CRITICAL",
        "Fixed in v5.0",
    ),
    (
        "2",
        "Description text area too small — no dedicated full-screen view",
        "User creates or views a bug. Description field appears in Bug Info area only.",
        "Dedicated full Description tab with large text editor (CC_DESC)",
        "Full editor (Screen 0320) not visible because tab crashes (Bug 1).",
        "Root cause = Bug 1 tab crash",
        "Fix Bug 1 -> full editor on tab Description (subscreen 0320) becomes accessible.",
        "—",
        "Medium",
        "Resolved by Bug 1 fix",
    ),
    (
        "3",
        "Description field character limit",
        "User types long description (>132 chars) in Description input area — text is cut off",
        "Unlimited text (TYPE STRING)",
        "Limited to ~132 visible characters",
        "STRING field GS_BUG_DETAIL-DESC_TEXT placed directly on screen layout -> screen field limits visible length",
        "Remove GS_BUG_DETAIL-DESC_TEXT from screen 0310 layout. Description stored only through cl_gui_textedit (CC_DESC_MINI).",
        "—",
        "Medium",
        "Fixed in v5.0",
    ),
    (
        "4",
        "SAP Module, Severity, Created Date fields display empty",
        "Open Display Bug for BUG0000023 (Screen 0310). SAP Module, Severity, Created Date all show blank.",
        "Fields display correct values from ZBUG_TRACKER",
        "Fields appear empty",
        "Screen 0310 fields not mapped to correct global variables GS_BUG_DETAIL-SAP_MODULE, SEVERITY, CREATED_AT; or fields not added to screen layout in SE51",
        "Verify SE51 Screen 0310 field references. Add missing fields to layout. Check SE16 data for BUG0000023.",
        "display-bug-bug0023-empty-metadata.png",
        "High",
        "Fix in SE51 screen 0310",
    ),
    (
        "5",
        "Remove User deletes row without selection",
        "In Project Detail Screen 0500, press Remove User without clicking any row. User is deleted anyway.",
        "Message: Please select a user row to remove. No deletion occurs.",
        "Row is silently removed (default current_line used)",
        "tc_users-current_line always has a value when TC has data. IF lv_line = 0 check insufficient.",
        "Replace with GET CURSOR LINE lv_line and validate lv_line > 0 AND lv_line <= lines(gt_user_project)",
        "change-project-remove-user-confirm.png",
        "Medium",
        "Fixed in v5.0 CODE_F01",
    ),
    (
        "6a",
        "Create Bug: Status field allows non-New value",
        "On Screen 0310 Create mode, open STATUS F4 help -> select status 3 (In Progress) -> save. Bug saved with status 3.",
        "Status must always be 1 (New) when creating a bug",
        "Status accepts any value from F4",
        "IF gs_bug_detail-status IS INITIAL guard bypassed when user explicitly selects via F4",
        "Force gs_bug_detail-status = gc_st_new unconditionally in save_bug. Lock STATUS field (screen group STS) in Create mode PBO.",
        "create-bug-bug-info-status-3-no-upload.png",
        "High",
        "Fixed in v5.0 CODE_F01 + CODE_PBO",
    ),
    (
        "6b",
        "Create Bug: SAP Module has no F4 search help",
        "On Screen 0310, click SAP Module field -> no F4 help popup",
        "F4 popup shows list: FI, MM, SD, ABAP, BASIS, PP, HR, QM",
        "No search help appears; field is free-text only",
        "Missing f4_sap_module FORM and POV module for SAP_MODULE field",
        "Add FORM f4_sap_module in CODE_F02 using F4IF_INT_TABLE_VALUE_REQUEST. Add POV module f4_bug_sapmodule in CODE_PAI.",
        "—",
        "High",
        "Fixed in v5.0 CODE_F02 + CODE_PAI",
    ),
    (
        "6c",
        "Create Bug: Cannot upload evidence before bug is saved",
        "On Screen 0310 Create mode, button UP_FILE (Upload Evidence) is hidden/disabled",
        "UP_FILE available in Create mode — auto-saves bug first, then uploads",
        "UP_FILE button not visible in Create mode",
        "UP_FILE excluded from Create mode exclusion list in PBO",
        "Remove UP_FILE from Create mode exclusion list. In upload_evidence, auto-call save_bug when gv_current_bug_id IS INITIAL AND gv_mode = gc_mode_create.",
        "—",
        "High",
        "Fixed in v5.0 CODE_PBO + CODE_F01",
    ),
    (
        "6d",
        "Create Bug: Created Date field shows empty before save",
        "On Screen 0310 Create mode, CREATED_AT field shows blank",
        "Created Date auto-populated with today date immediately",
        "Field blank until save is triggered",
        "CREATED_AT only set inside save_bug (post-save). PBO Create mode does not pre-fill.",
        "PBO Create mode: add gs_bug_detail-created_at = sy-datum and gs_bug_detail-created_time = sy-uzeit.",
        "—",
        "Medium",
        "Fixed in v5.0 CODE_PBO",
    ),
    (
        "7",
        "All screen fields lock after validation error",
        "On Screen 0310, enter invalid data -> press Save -> validation error triggers -> ALL fields become locked",
        "Error message in status bar, fields remain editable for correction",
        "All screen fields locked, user cannot correct input",
        "MESSAGE TYPE E in Module Pool locks screen fields for all non-grouped elements",
        "Replace all MESSAGE TYPE E in save_bug_detail and save_project_detail with MESSAGE TYPE S DISPLAY LIKE E followed by RETURN",
        "create-bug-validation-error-footer.png",
        "High",
        "Fixed in v5.0 CODE_F01",
    ),
    (
        "8",
        "Description text disappears in Display/Change mode",
        "Open existing bug in Display mode (Screen 0300) -> Bug Info tab -> mini Description area is empty",
        "Mini editor shows the bug description text",
        "Mini editor blank even when description was saved",
        "Description saved to Long Text Object (SAVE_TEXT) but not synced back to gs_bug_detail-desc_text DB field. Load re-reads DB field -> gets old/empty value.",
        "After save_long_text, sync back: read go_edit_desc text -> concatenate lines -> assign to gs_bug_detail-desc_text -> update DB",
        "—",
        "High",
        "Fixed in v5.0 CODE_F01",
    ),
    (
        "9",
        "Short dump CALL_FUNCTION_CONFLICT_TYPE in Change Bug",
        "Screen 0300 -> Change -> tab click -> short dump (same as Bug 1 but confirmed on Change path)",
        "Same as Bug 1",
        "CALL_FUNCTION_CONFLICT_TYPE short dump",
        "Same root cause as Bug 1",
        "Same fix as Bug 1",
        "runtime-error-read-text-type-conflict-123322.png",
        "CRITICAL",
        "Fixed in v5.0 (same as Bug 1)",
    ),
    (
        "10",
        "Manager can set status backward (3 -> 1) without warning",
        "Bug BUG0000024 status = 3 (In Progress). Manager opens Change -> changes status to 1 (New) -> saves. No warning.",
        "Transition 3->1 should be blocked. Status must follow allowed transition matrix.",
        "Status saved as 1 (New). System allows illegal backward transition.",
        "Manager role case in change_bug_status appends ALL statuses to lt_allowed — no backward transition check",
        "v5.0 redesign: Remove Manager bypass. Manager must follow transition matrix. Popup Screen 0370 only shows allowed next states.",
        "change-bug-bug0024-status-in-progress.png + change-bug-bug0024-status-1-new-saved.png",
        "CRITICAL",
        "Fixed in v5.0 CODE_F01 (validate_transition)",
    ),
    (
        "11",
        "Status transition to Fixed (5) allowed without evidence",
        "Bug BUG0000024 -> Change Status -> select Fixed (5) -> confirm -> saved without uploading any evidence file",
        "Transition to Fixed (5) requires at least 1 evidence file in ZBUG_EVIDENCE",
        "Status saved as Fixed with no evidence",
        "check_evidence_for_status logic bypassed by Manager role (IF sy-subrc <> 0 AND gv_role <> M)",
        "Remove Manager bypass in check_evidence_for_status. Popup Screen 0370 enforces evidence upload. ZBUG_EVIDENCE COUNT check applies to all roles.",
        "change-bug-bug0024-status-5-fixed-saved.png",
        "High",
        "Fixed in v5.0 CODE_F01",
    ),
]

for bug in bugs:
    f = sev_fill(bug[8])
    cl(ws, r, 1, bug[0], f, BOLD_FONT, CENTER)
    cl(ws, r, 2, bug[1], f, BOLD_FONT, LEFT)
    cl(ws, r, 3, bug[2], f, BODY_FONT, LEFT)
    cl(ws, r, 4, bug[3], f, BODY_FONT, LEFT)
    cl(ws, r, 5, bug[4], f, BODY_FONT, LEFT)
    cl(ws, r, 6, bug[5], f, BODY_FONT, LEFT)
    cl(ws, r, 7, bug[6], f, BODY_FONT, LEFT)
    cl(ws, r, 8, bug[7], f, BODY_FONT, LEFT)
    cl(ws, r, 9, bug[8], f, BOLD_FONT, CENTER)
    cl(ws, r, 10, bug[9], GREEN, BOLD_FONT, CENTER)
    ws.row_dimensions[r].height = 80
    r += 1

r += 1
r = section(ws, r, "Bug Statistics", 10)
hrow(ws, r, ["Severity", "Count", "Fixed in v5.0", "Pending", "", "", "", "", "", ""])
ws.merge_cells(f"E{r}:J{r}")
r += 1
for sev, cnt, fixed, pend in [
    ("CRITICAL", "3 (Bugs 1, 9, 10)", "3", "0"),
    ("High", "7", "7", "0"),
    ("Medium", "4 (Bugs 2, 3, 5, 6d)", "4", "0"),
]:
    f = sev_fill(sev)
    cl(ws, r, 1, sev, f, BOLD_FONT, CENTER)
    cl(ws, r, 2, cnt, f, BODY_FONT, CENTER)
    cl(ws, r, 3, fixed, GREEN, BODY_FONT, CENTER)
    cl(ws, r, 4, pend, f, BODY_FONT, CENTER)
    ws.merge_cells(f"E{r}:J{r}")
    r += 1
for c, v in enumerate(["Total", "11", "11", "0", "", "", "", "", "", ""], 1):
    cl(ws, r, c, v, HDR_FILL, HDR_FONT, CENTER)
ws.merge_cells(f"E{r}:J{r}")

widths(ws, [6, 30, 28, 28, 28, 32, 38, 30, 12, 20])
ws.freeze_panes = "A5"

# ============================================================
# Sheet 2: Issue 2 (Files Changed)
# ============================================================
ws2 = wb.create_sheet("Issue 2")
ws2.merge_cells("A1:F1")
ws2["A1"].value = "Files Changed in v5.0 Bug Fixes"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30

r = 3
hrow(
    ws2,
    r,
    ["SAP Include", "Code Reference", "Bugs Fixed", "Change Description", "", ""],
)
ws2.merge_cells(f"E{r}:F{r}")
r += 1
changed = [
    (
        "Z_BUG_WS_TOP",
        "CODE_TOP",
        "10, 11",
        "Add gc_st_finaltesting = 6, gc_st_resolved = V, screen group STS",
    ),
    (
        "Z_BUG_WS_PBO",
        "CODE_PBO",
        "1, 6a, 6c, 6d, 9",
        "TRY-CATCH for container creation; lock STATUS (group STS); UP_FILE in Create mode; pre-fill CREATED_AT",
    ),
    ("Z_BUG_WS_PAI", "CODE_PAI", "6b", "Add POV module f4_bug_sapmodule"),
    (
        "Z_BUG_WS_F01",
        "CODE_F01",
        "5, 7, 8, 10, 11",
        "Fix remove_user_from_project cursor check; replace MESSAGE TYPE E with TYPE S DISPLAY LIKE E; sync desc_text after save_long_text; enforce transition matrix for Manager; remove Manager bypass in evidence check",
    ),
    (
        "Z_BUG_WS_F02",
        "CODE_F02",
        "1, 6b, 9",
        "Explicit type cast lv_tdname TYPE tdobname before READ_TEXT; add f4_sap_module FORM",
    ),
]
for rd in changed:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws2, r, 1, rd[0], f, BOLD_FONT, LEFT)
    cl(ws2, r, 2, rd[1], f, BODY_FONT, CENTER)
    cl(ws2, r, 3, rd[2], f, BODY_FONT, CENTER)
    ws2.merge_cells(f"D{r}:F{r}")
    cl(ws2, r, 4, rd[3], f, BODY_FONT, LEFT)
    ws2.row_dimensions[r].height = 45
    r += 1

r += 1
r = section(ws2, r, "Screens to Verify in SE51", 6)
hrow(ws2, r, ["Screen", "Action Required", "Bugs Addressed", "", "", ""])
ws2.merge_cells(f"D{r}:F{r}")
r += 1
screens = [
    (
        "0310",
        "Remove STRING fields from layout; add SAP_MODULE, SEVERITY, CREATED_AT; add STATUS to group STS; add POV for SAP_MODULE",
        "3, 4, 6a, 6b, 6d",
    ),
    ("0320", "Verify Custom Control CC_DESC exists", "1, 2, 9"),
    ("0330", "Verify Custom Control CC_DEVNOTE exists (no underscore)", "1, 9"),
    ("0340", "Verify Custom Control CC_TSTRNOTE exists (no underscore)", "1, 9"),
]
for rd in screens:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws2, r, 1, rd[0], f, BOLD_FONT, CENTER)
    ws2.merge_cells(f"B{r}:C{r}")
    cl(ws2, r, 2, rd[1], f, BODY_FONT, LEFT)
    ws2.merge_cells(f"D{r}:F{r}")
    cl(ws2, r, 4, rd[2], f, BODY_FONT, CENTER)
    ws2.row_dimensions[r].height = 35
    r += 1

widths(ws2, [20, 16, 16, 42, 12, 12])
ws2.freeze_panes = "A4"

# ============================================================
# Sheet 3: Issue 4 (UAT Round 2 Regression Plan)
# ============================================================
ws3 = wb.create_sheet("Issue 4")
ws3.merge_cells("A1:F1")
ws3["A1"].value = "UAT Round 2 — Regression Plan"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

r = 3
hrow(ws3, r, ["Re-test Item", "Bug Ref", "Test Account", "Pass/Fail", "Date", "Notes"])
r += 1
regression = [
    (
        "Open Change Bug -> tab Description/Dev Note/Tester Note",
        "1, 9",
        "DEV-089",
        "",
        "",
        "",
    ),
    ("Description full editor shows content", "2, 3, 8", "DEV-089", "", "", ""),
    (
        "SAP Module, Severity, Created Date visible on Bug Info",
        "4",
        "DEV-089",
        "",
        "",
        "",
    ),
    ("Remove User requires row selection", "5", "DEV-089", "", "", ""),
    ("Create Bug -> status forced to New", "6a", "DEV-089", "", "", ""),
    ("Create Bug -> SAP Module F4 shows list", "6b", "DEV-089", "", "", ""),
    (
        "Create Bug -> UP_FILE button visible + auto-saves bug",
        "6c",
        "DEV-089",
        "",
        "",
        "",
    ),
    ("Create Bug -> Created Date pre-filled", "6d", "DEV-089", "", "", ""),
    ("Validation error -> fields remain editable", "7", "DEV-089", "", "", ""),
    ("Manager cannot set status backward", "10", "DEV-089", "", "", ""),
    ("Fixed requires evidence upload", "11", "DEV-089", "", "", ""),
]
for rd in regression:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws3, r, c, v, f, BODY_FONT, CENTER if c in (2, 3, 4, 5) else LEFT)
    ws3.row_dimensions[r].height = 25
    r += 1

r += 2
r = section(ws3, r, "Issue Category Breakdown", 6)
hrow(ws3, r, ["Category", "Bugs", "Description", "", "", ""])
ws3.merge_cells(f"D{r}:F{r}")
r += 1
for rd in [
    (
        "A: Screen Layout / SE51 Issues",
        "1, 2, 3, 4, 9",
        "All require SE51 screen layout verification and correction.",
    ),
    ("B: Business Logic (ABAP Code)", "5, 7, 8, 10, 11", "Fixed in CODE_F01 v5.0."),
    (
        "C: Missing Features (Field / F4 Help)",
        "6a, 6b, 6c, 6d",
        "Missing validations and helpers added in CODE_PAI + CODE_F02 v5.0.",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws3, r, 1, rd[0], f, BOLD_FONT, LEFT)
    cl(ws3, r, 2, rd[1], f, BODY_FONT, CENTER)
    ws3.merge_cells(f"C{r}:F{r}")
    cl(ws3, r, 3, rd[2], f, BODY_FONT, LEFT)
    ws3.row_dimensions[r].height = 25
    r += 1

widths(ws3, [46, 16, 14, 12, 14, 20])
ws3.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
