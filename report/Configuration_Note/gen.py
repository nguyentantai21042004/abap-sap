import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Configuration_Note.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)


def drow(ws, r, vals, fill=None):
    f = ALT_FILL if r % 2 == 0 else None
    if fill:
        f = fill
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, span="A:H"):
    cols = span.split(":")
    c1 = cols[0]
    c2 = cols[1] if len(cols) > 1 else cols[0]
    ws.merge_cells(f"{c1}{r}:{c2}{r}")
    cl(ws, r, ord(c1) - 64, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
x = ws["A1"]
x.value = "Configuration Note"
x.fill = HDR_FILL
x.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
x.alignment = CENTER
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
x2 = ws["A2"]
x2.value = "Z_BUG_WORKSPACE_MP — Bug Tracking System"
x2.fill = SUB_FILL
x2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
x2.alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
info = [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Module", "ABAP / Cross-Module (FI, MM, SD, BASIS)"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("Package", "ZBUGTRACK"),
    ("T-Code", "ZBUG_WS"),
    ("Program Type", "Module Pool (Type M)"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
    ("Prepared by", "DEV-089 (Manager)"),
]
for k, v in info:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

widths(ws, [20, 30, 15, 15, 15, 15, 15, 15])
ws.freeze_panes = "A4"

# ============================================================
# Sheet 2: Record of change
# ============================================================
ws2 = wb.create_sheet("Record of change")
ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "Record of Change — Configuration Note"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws2.row_dimensions[1].height = 30

r = 3
hrow(
    ws2,
    r,
    ["No.", "Effective Date", "Version", "Change Description", "Changed by", "Remarks"],
)
r += 1
drow(
    ws2,
    r,
    ["1", "17/04/2026", "1.0", "Initial configuration note for v5.0", "DEV-089", ""],
)
widths(ws2, [6, 16, 10, 55, 14, 20])
ws2.freeze_panes = "A4"

# ============================================================
# Sheet 3: Checklist
# ============================================================
ws3 = wb.create_sheet("Checklist")
ws3.merge_cells("A1:F1")
t = ws3["A1"]
t.value = "Configuration Checklist — Z_BUG_WORKSPACE_MP v5.0"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws3.row_dimensions[1].height = 30

r = 3
r = section(ws3, r, "SE11 — ABAP Dictionary Objects", "A:F")
hrow(ws3, r, ["#", "Object Type", "Object Name", "Description", "Status", "Remarks"])
r += 1
se11 = [
    (
        "1",
        "Domain",
        "zde_bug_status",
        "Bug status codes (CHAR 20): 1/2/3/4/5/6/7/W/R/V",
        "Done",
        "",
    ),
    (
        "2",
        "Domain",
        "zde_sap_module",
        "SAP module codes (CHAR 20): FI/MM/SD/ABAP/BASIS",
        "Done",
        "",
    ),
    ("3", "Domain", "zde_bug_role", "User roles (CHAR 1): M/D/T", "Done", ""),
    ("4", "Data Element", "ZDE_BUG_ID", "Bug ID (CHAR 10)", "Done", ""),
    ("5", "Data Element", "ZDE_PROJECT_ID", "Project ID (CHAR 20)", "Done", ""),
    ("6", "Data Element", "ZDE_PRIORITY", "Priority (CHAR 1): H/M/L", "Done", ""),
    ("7", "Data Element", "ZDE_SEVERITY", "Severity (CHAR 1): D/V/H/N/M", "Done", ""),
    ("8", "Data Element", "ZDE_BUG_TYPE", "Bug type (CHAR 1)", "Done", ""),
    ("9", "Data Element", "ZDE_REASONS", "Root causes (STRING)", "Done", ""),
    ("10", "Data Element", "ZDE_USERNAME", "SAP username (CHAR 12)", "Done", ""),
    ("11", "Data Element", "ZDE_IS_DEL", "Soft-delete flag (CHAR 1)", "Done", ""),
    ("12", "Table", "ZBUG_TRACKER", "Core bug records — 29 fields", "Done", ""),
    ("13", "Table", "ZBUG_USERS", "User registry — 12 fields", "Done", ""),
    ("14", "Table", "ZBUG_PROJECT", "Project records — 16 fields", "Done", ""),
    (
        "15",
        "Table",
        "ZBUG_USER_PROJEC",
        "User-Project M:N mapping — 10 fields",
        "Done",
        "",
    ),
    ("16", "Table", "ZBUG_HISTORY", "Bug change history — 10 fields", "Done", ""),
    (
        "17",
        "Table",
        "ZBUG_EVIDENCE",
        "Binary evidence storage — 11 fields",
        "Pending",
        "Create via TR-01",
    ),
    ("18", "Package", "ZBUGTRACK", "Development package for all objects", "Done", ""),
]
for rd in se11:
    f = GREEN if rd[4] == "Done" else (YELLOW if rd[4] == "Pending" else None)
    drow(ws3, r, rd)
    cl(ws3, r, 5, rd[4], f, BODY_FONT, CENTER)
    r += 1

r += 1
r = section(ws3, r, "SE38 — ABAP Program Objects", "A:F")
hrow(ws3, r, ["#", "Object Name", "Type", "Description", "Status", "Remarks"])
r += 1
se38 = [
    (
        "1",
        "Z_BUG_WORKSPACE_MP",
        "Module Pool (M)",
        "Main program",
        "Done",
        "v5.0 code ready",
    ),
    (
        "2",
        "Z_BUG_WS_TOP",
        "Include",
        "Global declarations, types, constants",
        "Done",
        "",
    ),
    (
        "3",
        "Z_BUG_WS_F00",
        "Include",
        "ALV field catalog + LCL_EVENT_HANDLER",
        "Done",
        "",
    ),
    ("4", "Z_BUG_WS_PBO", "Include", "Process Before Output modules", "Done", ""),
    ("5", "Z_BUG_WS_PAI", "Include", "Process After Input modules", "Done", ""),
    ("6", "Z_BUG_WS_F01", "Include", "Business logic FORMs", "Done", ""),
    (
        "7",
        "Z_BUG_WS_F02",
        "Include",
        "Helpers: F4, Long Text, Popup, Download",
        "Done",
        "",
    ),
]
for rd in se38:
    drow(ws3, r, rd)
    cl(ws3, r, 5, rd[4], GREEN, BODY_FONT, CENTER)
    r += 1

r += 1
r = section(ws3, r, "SE51 — Screens", "A:F")
hrow(ws3, r, ["#", "Screen", "Type", "Description", "Status", "Remarks"])
r += 1
se51 = [
    ("1", "0200", "Normal", "Bug List (ALV + Dashboard header)", "Done", ""),
    ("2", "0210", "Modal Dialog", "Bug Search Input popup", "Pending", "F11"),
    ("3", "0220", "Normal", "Bug Search Results ALV", "Pending", "F11"),
    ("4", "0300", "Normal + Tab Strip", "Bug Detail (6 subscreens)", "Done", ""),
    (
        "5",
        "0310-0360",
        "Subscreens",
        "Bug Info / Description / Dev Note / Tester Note / Evidence / History",
        "Done",
        "",
    ),
    ("6", "0370", "Modal Dialog", "Status Transition popup", "Pending", "F11"),
    ("7", "0400", "Normal", "Project List ALV", "Done", ""),
    ("8", "0410", "Normal", "Project Search (initial screen)", "Pending", "F11"),
    (
        "9",
        "0500",
        "Normal + Table Control",
        "Project Detail + User Assignment",
        "Done",
        "",
    ),
]
for rd in se51:
    f = GREEN if rd[4] == "Done" else (YELLOW if rd[4] == "Pending" else None)
    drow(ws3, r, rd)
    cl(ws3, r, 5, rd[4], f, BODY_FONT, CENTER)
    r += 1

r += 1
r = section(ws3, r, "SE93 / SE91 / SMW0 / SE75", "A:F")
hrow(ws3, r, ["#", "Transaction", "Object Name", "Description", "Status", "Remarks"])
r += 1
misc = [
    (
        "1",
        "SE93",
        "ZBUG_WS",
        "T-Code -> Z_BUG_WORKSPACE_MP, initial screen 0410",
        "Update needed",
        "Change 0400->0410",
    ),
    ("2", "SE91", "ZBUG_MSG", "Message class — 33 messages (EN + VI)", "Done", ""),
    (
        "3",
        "SMW0",
        "ZBT_TMPL_01",
        "Bug_report.xlsx — Tester uploads bug report",
        "Done",
        "",
    ),
    (
        "4",
        "SMW0",
        "ZBT_TMPL_02",
        "fix_report.xlsx — Developer uploads fix evidence",
        "Done",
        "",
    ),
    (
        "5",
        "SMW0",
        "ZBT_TMPL_03",
        "confirm_report.xlsx — Final Tester confirmation",
        "Done",
        "",
    ),
    (
        "6",
        "SMW0",
        "ZTEMPLATE_PROJECT",
        "project_template.xlsx — Manager bulk upload",
        "Done",
        "",
    ),
    ("7", "SE75", "ZBUG_NOTE / Z001", "Bug Description long text", "Done", ""),
    ("8", "SE75", "ZBUG_NOTE / Z002", "Developer Note long text", "Done", ""),
    ("9", "SE75", "ZBUG_NOTE / Z003", "Tester Note long text", "Done", ""),
]
for rd in misc:
    f = (
        GREEN
        if rd[4] == "Done"
        else (YELLOW if "Pending" in rd[4] or "Update" in rd[4] else None)
    )
    drow(ws3, r, rd)
    cl(ws3, r, 5, rd[4], f, BODY_FONT, CENTER)
    r += 1

widths(ws3, [5, 22, 20, 50, 14, 22])
ws3.freeze_panes = "A4"

# ============================================================
# Sheet 4: Customizing Guide (SE11)
# ============================================================
ws4 = wb.create_sheet("Customizing Guide (SE11)")
ws4.merge_cells("A1:G1")
t = ws4["A1"]
t.value = "Customizing Guide — SE11 ABAP Dictionary"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
r = section(ws4, r, "Status Values — Breaking Change v4.x to v5.0", "A:G")
hrow(
    ws4,
    r,
    ["Code", "Meaning", "ABAP Constant", "In v4.x", "In v5.0", "Direction", "Note"],
)
r += 1
statuses = [
    ("1", "New", "gc_st_new", "New", "New", "Same", ""),
    ("W", "Waiting", "gc_st_waiting", "Waiting", "Waiting", "Same", ""),
    ("2", "Assigned", "gc_st_assigned", "Assigned", "Assigned", "Same", ""),
    ("3", "In Progress", "gc_st_inprogress", "In Progress", "In Progress", "Same", ""),
    ("4", "Pending", "gc_st_pending", "Pending", "Pending", "Same", ""),
    ("5", "Fixed", "gc_st_fixed", "Fixed", "Fixed", "Same", ""),
    ("R", "Rejected", "gc_st_rejected", "Rejected", "Rejected", "Same", ""),
    (
        "6",
        "CHANGED",
        "gc_st_finaltesting",
        "Resolved",
        "Final Testing",
        "BREAKING",
        "6 was Resolved in v4.x",
    ),
    (
        "V",
        "Resolved",
        "gc_st_resolved",
        "(did not exist)",
        "Resolved (terminal)",
        "NEW",
        "New in v5.0",
    ),
    ("7", "Closed (legacy)", "gc_st_closed", "Closed", "Closed", "Same", "Legacy only"),
]
for rd in statuses:
    f = RED if rd[5] == "BREAKING" else (GREEN if rd[5] == "NEW" else None)
    for c, v in enumerate(rd, 1):
        cl(
            ws4,
            r,
            c,
            v,
            f,
            BODY_FONT if rd[5] not in ("BREAKING", "NEW") else BOLD_FONT,
            LEFT,
        )
    r += 1

r += 1
r = section(ws4, r, "ZBUG_TRACKER — Key Field Types", "A:G")
hrow(ws4, r, ["Field", "Data Type", "Length", "ABAP Domain", "Note", "", ""])
ws4.merge_cells(f"F{r}:G{r}")
r += 1
fields = [
    ("STATUS", "CHAR", "20", "zde_bug_status", "NOT CHAR 1 — domain has 10 values"),
    (
        "SAP_MODULE",
        "CHAR",
        "20",
        "zde_sap_module",
        "NOT CHAR 10 — domain has module codes",
    ),
    ("DESC_TEXT", "STRING", "—", "—", "Cannot place directly on Dynpro layout"),
    ("REASONS", "STRING", "—", "—", "Cannot place directly on Dynpro layout"),
]
for rd in fields:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, LEFT)
    ws4.merge_cells(f"F{r}:G{r}")
    r += 1

r += 1
r = section(ws4, r, "Screen Groups — ZBUG_TRACKER Fields", "A:G")
hrow(ws4, r, ["Group", "Fields Included", "Behaviour", "", "", "", ""])
ws4.merge_cells(f"D{r}:G{r}")
r += 1
groups = [
    ("EDT", "All editable fields", "Locked in Display mode"),
    ("BID", "BUG_ID", "Always locked — auto-generated"),
    ("PRJ", "PROJECT_ID", "Locked when coming from project context"),
    ("FNC", "BUG_TYPE, PRIORITY, SEVERITY", "Locked for Developer role"),
    ("TST", "TESTER_ID", "Locked for Developer role"),
    ("DEV", "DEV_ID, VERIFY_TESTER_ID", "Locked for Tester role"),
    ("STS", "STATUS", "Always locked — change via popup 0370 only"),
]
for rd in groups:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws4, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws4, r, 2, rd[1], f, BODY_FONT, LEFT)
    ws4.merge_cells(f"C{r}:G{r}")
    cl(ws4, r, 3, rd[2], f, BODY_FONT, LEFT)
    r += 1

widths(ws4, [12, 28, 12, 18, 38, 12, 12])
ws4.freeze_panes = "A4"

# ============================================================
# Sheet 5: Customizing Guide (SE51)
# ============================================================
ws5 = wb.create_sheet("Customizing Guide (SE51)")
ws5.merge_cells("A1:G1")
t = ws5["A1"]
t.value = "Customizing Guide — SE51 Screen Painter"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws5.row_dimensions[1].height = 30

r = 3
r = section(ws5, r, "Pending Screens to Create (F11-F16)", "A:G")
hrow(ws5, r, ["Step", "Action", "SAP Transaction", "Priority", "Status", "Notes", ""])
ws5.merge_cells(f"G{r}:{get_column_letter(7)}{r}")
r += 1
pending = [
    (
        "F11",
        "Create screens 0410, 0370, 0210, 0220",
        "SE51",
        "HIGH",
        "TODO",
        "Initial screen + popups",
    ),
    (
        "F12",
        "Create GUI Statuses: STATUS_0410, STATUS_0370, STATUS_0210, STATUS_0220",
        "SE41",
        "HIGH",
        "TODO",
        "",
    ),
    (
        "F12",
        "Update STATUS_0200: add SEARCH button (FCode = SEARCH)",
        "SE41",
        "HIGH",
        "TODO",
        "",
    ),
    (
        "F13",
        "Update ZBUG_WS: change initial screen 0400 -> 0410",
        "SE93",
        "MEDIUM",
        "TODO",
        "",
    ),
    (
        "F14",
        "Copy all v5.0 code includes into SAP",
        "SE38/SE80",
        "HIGH",
        "TODO",
        "TR-07",
    ),
    ("F15", "Create ZBUG_EVIDENCE table", "SE11", "HIGH", "TODO", "TR-01"),
    (
        "F16",
        "Run status migration script (6 -> V)",
        "SE38 / SE16N",
        "HIGH",
        "TODO",
        "Run once after deploy",
    ),
]
for rd in pending:
    f = RED if rd[3] == "HIGH" else YELLOW
    for c, v in enumerate(rd, 1):
        cl(ws5, r, c, v, f, BODY_FONT, LEFT)
    ws5.merge_cells(f"G{r}:{get_column_letter(7)}{r}")
    r += 1

r += 1
r = section(ws5, r, "Screen 0300 — Tab Strip Configuration", "A:G")
hrow(
    ws5,
    r,
    [
        "Tab No.",
        "Tab Title",
        "Subscreen",
        "Custom Control",
        "Long Text Object",
        "Text ID",
        "Notes",
    ],
)
r += 1
tabs = [
    ("1", "Bug Info", "0310", "—", "—", "—", "Main fields, no text editor"),
    ("2", "Description", "0320", "CC_DESC", "ZBUG_NOTE", "Z001", "Full text editor"),
    ("3", "Dev Note", "0330", "CC_DEVNOTE", "ZBUG_NOTE", "Z002", "Developer notes"),
    ("4", "Tester Note", "0340", "CC_TSTRNOTE", "ZBUG_NOTE", "Z003", "Tester notes"),
    ("5", "Evidence", "0350", "CC_EVD_LIST", "—", "—", "ALV grid for files"),
    ("6", "History", "0360", "CC_HIST_LIST", "—", "—", "ALV grid for history"),
]
for rd in tabs:
    drow(ws5, r, rd)
    r += 1

r += 1
r = section(ws5, r, "GUI Status Buttons Reference", "A:G")
hrow(ws5, r, ["Screen", "GUI Status", "Key Buttons (FCode)", "Type", "Notes", "", ""])
ws5.merge_cells(f"F{r}:G{r}")
r += 1
statuses_gui = [
    (
        "0200",
        "STATUS_0200",
        "CREATE, CHANGE, DISPLAY, DELETE, REFRESH, DN_TC, DN_PROOF, DN_CONF, SEARCH, MY_BUGS",
        "Normal",
        "Update: add SEARCH",
    ),
    (
        "0210",
        "STATUS_0210",
        "EXECUTE (F8), CANCEL (F12)",
        "Modal Dialog",
        "New in v5.0",
    ),
    (
        "0220",
        "STATUS_0220",
        "BACK (F3), EXIT (Shift+F3), CANCEL (F12)",
        "Normal",
        "New in v5.0",
    ),
    (
        "0300",
        "STATUS_0300",
        "SAVE, BACK, CANCEL, STATUS_CHG, UP_FILE, UP_REPORT, UP_FIX, DN_EVD, SENDMAIL",
        "Normal",
        "",
    ),
    (
        "0370",
        "STATUS_0370",
        "CONFIRM, UP_TRANS, CANCEL (F12)",
        "Modal Dialog",
        "New in v5.0",
    ),
    (
        "0400",
        "STATUS_0400",
        "CREATE_PRJ, CHANGE, DISPLAY, DELETE, MY_BUGS, DN_PRJ, UP_PRJ, REFRESH",
        "Normal",
        "",
    ),
    (
        "0410",
        "STATUS_0410",
        "EXECUTE (F8), BACK (F3), EXIT (Shift+F3), CANCEL (F12)",
        "Normal",
        "New in v5.0",
    ),
    ("0500", "STATUS_0500", "SAVE, BACK, CANCEL, ADD_USER, REMOVE_USER", "Normal", ""),
]
for rd in statuses_gui:
    f = YELLOW if "New in v5.0" in rd[4] else (ALT_FILL if r % 2 == 0 else None)
    for c, v in enumerate(rd[:5], 1):
        cl(ws5, r, c, v, f, BODY_FONT, LEFT)
    ws5.merge_cells(f"F{r}:G{r}")
    r += 1

widths(ws5, [8, 20, 55, 16, 22, 12, 12])
ws5.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
