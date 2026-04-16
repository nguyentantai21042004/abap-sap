import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Functional_Specification.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=8):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
ws["A1"].value = "Functional Specification"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 40
ws.merge_cells("A2:H2")
ws["A2"].value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Function ID", "ZBUG_WS_v5"),
    ("Module", "ABAP Cross-Module Bug Tracking"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Program", "Z_BUG_WORKSPACE_MP (Module Pool, Type M)"),
    ("Created by", "DEV-089"),
    ("Reviewed by", "—"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1
widths(ws, [20, 32, 12, 12, 12, 12, 12, 12])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
ws2["A1"].value = "Change History — Functional Specification"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30
r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2,
    r,
    ["1", "1.0", "Initial functional spec — v5.0", "All", "17/04/2026", "DEV-089", ""],
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: Function Overview
# ============================================================
ws3 = wb.create_sheet("Function Overview")
ws3.merge_cells("A1:F1")
ws3["A1"].value = "Function Overview"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30
r = 3
for k, v in [
    ("Function ID", "ZBUG_WS_v5"),
    ("Processing Time", "Online (real-time Dynpro interaction)"),
    ("Processing Type", "Multilingual (EN/VI)"),
    (
        "Function Overview",
        "Custom ABAP Module Pool for centralised bug tracking across SAP development projects. Supports 3 roles (Manager/Developer/Tester), 10-state bug lifecycle, auto-assignment, evidence management, email notifications, and project-based access control.",
    ),
    (
        "Supplement",
        "Replaces manual Excel-based bug tracking. Integrates with SAP BCS email API and SMW0 file templates.",
    ),
]:
    cl(ws3, r, 1, k, None, BOLD_FONT, LEFT)
    ws3.merge_cells(f"B{r}:F{r}")
    cl(ws3, r, 2, v, None, BODY_FONT, LEFT)
    ws3.row_dimensions[r].height = 35 if len(v) > 80 else 20
    r += 1
widths(ws3, [20, 60, 12, 12, 12, 12])

# ============================================================
# Sheet 4: Process Flow
# ============================================================
ws4 = wb.create_sheet("Process Flow")
ws4.merge_cells("A1:F1")
ws4["A1"].value = "Business Process Flow"
ws4["A1"].fill = HDR_FILL
ws4["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
r = section(ws4, r, "Overall System Flow (Screen Navigation)", 6)
hrow(ws4, r, ["Step", "From Screen", "Action", "To Screen", "Description", "Role"])
r += 1
flow = [
    (
        "1",
        "SAP GUI",
        "Run /nZBUG_WS",
        "0410 — Project Search",
        "Initial screen — 3 filter fields: Project ID, Manager, Status",
        "All",
    ),
    (
        "2",
        "0410",
        "Execute (F8)",
        "0400 — Project List",
        "ALV showing filtered projects",
        "All",
    ),
    (
        "3",
        "0400",
        "Create Project",
        "0500 — Project Detail",
        "New project + user assignment",
        "M only",
    ),
    (
        "4",
        "0400",
        "Change/Display Project",
        "0500 — Project Detail",
        "Edit existing project",
        "M only",
    ),
    (
        "5",
        "0400",
        "Double-click project",
        "0200 — Bug List",
        "ALV with Dashboard header showing status counts",
        "All",
    ),
    (
        "6",
        "0200",
        "Create Bug",
        "0300 — Bug Detail",
        "6-tab screen: Bug Info/Desc/Dev Note/Tester Note/Evidence/History",
        "M, T",
    ),
    (
        "7",
        "0300",
        "Change Status",
        "0370 — Status Transition Popup",
        "Modal dialog: select new status + notes",
        "All (per matrix)",
    ),
    (
        "8",
        "0200",
        "SEARCH button",
        "0210 — Bug Search Popup",
        "6 filter fields: Bug ID, Title, Status, Priority, Module, Reporter",
        "All",
    ),
    (
        "9",
        "0210",
        "Execute (F8)",
        "0220 — Search Results ALV",
        "Results filtered by criteria",
        "All",
    ),
    (
        "10",
        "0400",
        "My Bugs",
        "0200 — Bug List (filtered)",
        "Shows only bugs assigned to current user",
        "D, T",
    ),
]
for rd in flow:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, CENTER if c in (1, 4, 6) else LEFT)
    ws4.row_dimensions[r].height = 30
    r += 1

r += 1
r = section(ws4, r, "Bug Lifecycle Flow (Status Transitions)", 6)
hrow(
    ws4,
    r,
    [
        "From Status",
        "To Status",
        "Who Can Trigger",
        "Condition",
        "Auto-Action",
        "Notes",
    ],
)
r += 1
lifecycle = [
    (
        "New (1)",
        "Assigned (2)",
        "Manager (M)",
        "Developer assigned manually",
        "DEV_ID set, email sent",
        "Via popup 0370",
    ),
    (
        "New (1)",
        "Waiting (W)",
        "Auto",
        "No Developer available for module",
        "Email sent to Manager",
        "Auto-assign result",
    ),
    (
        "Waiting (W)",
        "Assigned (2)",
        "Manager (M)",
        "Manager manually assigns Dev",
        "DEV_ID set",
        "Via popup 0370",
    ),
    (
        "Assigned (2)",
        "In Progress (3)",
        "Developer (D)",
        "Dev accepts",
        "—",
        "Via popup 0370",
    ),
    (
        "Assigned (2)",
        "Rejected (R)",
        "Developer (D)",
        "Dev rejects — TRANS_NOTE required",
        "—",
        "Via popup 0370",
    ),
    (
        "In Progress (3)",
        "Fixed (5)",
        "Developer (D)",
        "Evidence required in ZBUG_EVIDENCE",
        "Auto-assign Tester -> FinalTesting",
        "Via popup 0370",
    ),
    (
        "In Progress (3)",
        "Pending (4)",
        "Developer (D)",
        "Dev pauses work",
        "—",
        "Via popup 0370",
    ),
    (
        "In Progress (3)",
        "Rejected (R)",
        "Developer (D)",
        "Dev rejects — TRANS_NOTE required",
        "—",
        "Via popup 0370",
    ),
    (
        "Pending (4)",
        "Assigned (2)",
        "Manager (M)",
        "Manager reassigns",
        "DEV_ID updated",
        "Via popup 0370",
    ),
    (
        "Fixed (5)",
        "Final Testing (6)",
        "Auto",
        "Tester auto-assigned",
        "VERIFY_TESTER_ID set, email sent",
        "After Dev sets Fixed",
    ),
    (
        "Fixed (5)",
        "Waiting (W)",
        "Auto",
        "No Tester available",
        "Email to Manager",
        "Auto-assign result",
    ),
    (
        "Final Testing (6)",
        "Resolved (V)",
        "Tester (T)",
        "TRANS_NOTE required — TERMINAL",
        "—",
        "Via popup 0370",
    ),
    (
        "Final Testing (6)",
        "In Progress (3)",
        "Tester (T)",
        "Test failed — TRANS_NOTE required",
        "—",
        "Via popup 0370",
    ),
]
for rd in lifecycle:
    f = GREEN if "Resolved" in rd[1] else (ALT_FILL if r % 2 == 0 else None)
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, LEFT)
    ws4.row_dimensions[r].height = 25
    r += 1

widths(ws4, [18, 18, 20, 36, 30, 18])
ws4.freeze_panes = "A4"

# ============================================================
# Sheet 5: Screen Layout
# ============================================================
ws5 = wb.create_sheet("Screen Layout")
ws5.merge_cells("A1:G1")
ws5["A1"].value = "Screen Layout Summary"
ws5["A1"].fill = HDR_FILL
ws5["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws5["A1"].alignment = CENTER
ws5.row_dimensions[1].height = 30

r = 3
hrow(
    ws5,
    r,
    [
        "Screen",
        "Type",
        "Size",
        "GUI Status",
        "Title Bar",
        "Description",
        "Initial from",
    ],
)
r += 1
screens = [
    (
        "0410",
        "Normal",
        "800x600",
        "STATUS_0410",
        "—",
        "Project Search — initial screen (v5.0)",
        "ZBUG_WS t-code",
    ),
    (
        "0400",
        "Normal",
        "800x600",
        "STATUS_0400",
        "—",
        "Project List ALV",
        "0410 Execute",
    ),
    (
        "0500",
        "Normal + Table Ctrl",
        "800x600",
        "STATUS_0500",
        "—",
        "Project Detail + User Assignment (Table Control TC_USERS)",
        "0400 Create/Change",
    ),
    (
        "0200",
        "Normal",
        "800x600",
        "STATUS_0200",
        "—",
        "Bug List ALV + Dashboard header",
        "0400 double-click",
    ),
    (
        "0210",
        "Modal Dialog",
        "800x400",
        "STATUS_0210",
        "T_0210 Bug Search",
        "Bug Search Input — 6 filter fields",
        "0200 SEARCH",
    ),
    (
        "0220",
        "Normal",
        "800x600",
        "STATUS_0220",
        "T_0220 Search Results",
        "Bug Search Results ALV",
        "0210 Execute",
    ),
    (
        "0300",
        "Normal + Tab Strip",
        "800x600",
        "STATUS_0300",
        "—",
        "Bug Detail — 6 tabs (subscreens 0310-0360)",
        "0200 Create/Change",
    ),
    (
        "0310",
        "Subscreen",
        "—",
        "—",
        "—",
        "Bug Info — main fields, mini description editor",
        "Tab 1 of 0300",
    ),
    (
        "0320",
        "Subscreen",
        "—",
        "—",
        "—",
        "Description — full CC_DESC text editor (ZBUG_NOTE Z001)",
        "Tab 2 of 0300",
    ),
    (
        "0330",
        "Subscreen",
        "—",
        "—",
        "—",
        "Dev Note — CC_DEVNOTE text editor (ZBUG_NOTE Z002)",
        "Tab 3 of 0300",
    ),
    (
        "0340",
        "Subscreen",
        "—",
        "—",
        "—",
        "Tester Note — CC_TSTRNOTE text editor (ZBUG_NOTE Z003)",
        "Tab 4 of 0300",
    ),
    (
        "0350",
        "Subscreen",
        "—",
        "—",
        "—",
        "Evidence — ALV list: EVD_ID, FILE_NAME, MIME_TYPE, FILE_SIZE, UPLOADED_BY, UPLOAD_DATE",
        "Tab 5 of 0300",
    ),
    (
        "0360",
        "Subscreen",
        "—",
        "—",
        "—",
        "History — ALV: LOG_ID, ACTION_TYPE, CHANGED_BY, CHANGED_AT, OLD_VALUE, NEW_VALUE, REASON",
        "Tab 6 of 0300",
    ),
    (
        "0370",
        "Modal Dialog",
        "600x400",
        "STATUS_0370",
        "T_0370 Change Status",
        "Status Transition popup — dropdown + conditional fields",
        "0300 Change Status",
    ),
]
for rd in screens:
    f = (
        YELLOW
        if "new" in rd[6].lower()
        or "0410" in rd[0]
        or "0370" in rd[0]
        or "0210" in rd[0]
        or "0220" in rd[0]
        else (ALT_FILL if r % 2 == 0 else None)
    )
    for c, v in enumerate(rd, 1):
        cl(ws5, r, c, v, f, BODY_FONT, LEFT)
    ws5.row_dimensions[r].height = 30
    r += 1

widths(ws5, [8, 18, 12, 18, 22, 46, 20])
ws5.freeze_panes = "A4"

# ============================================================
# Sheet 6: Screen Definition
# ============================================================
ws6 = wb.create_sheet("Screen Definition")
ws6.merge_cells("A1:G1")
ws6["A1"].value = "Screen Field Definition"
ws6["A1"].fill = HDR_FILL
ws6["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws6["A1"].alignment = CENTER
ws6.row_dimensions[1].height = 30

r = 3
r = section(ws6, r, "Screen 0310 — Bug Info (Tab 1)", 7)
hrow(
    ws6,
    r,
    ["No.", "Field Name", "Field Label", "Type", "Mandatory", "Screen Group", "Notes"],
)
r += 1
for rd in [
    ("1", "BUG_ID", "Bug ID", "CHAR 10", "Auto", "BID", "Always locked"),
    (
        "2",
        "PROJECT_ID",
        "Project ID",
        "CHAR 20",
        "Auto",
        "PRJ",
        "Locked from project context",
    ),
    ("3", "TITLE", "Title", "CHAR 100", "Y", "EDT", "Editable by all roles"),
    (
        "4",
        "SAP_MODULE",
        "SAP Module",
        "CHAR 20 (domain)",
        "Y",
        "FNC",
        "Locked for Developer",
    ),
    ("5", "BUG_TYPE", "Bug Type", "CHAR 1", "Y", "FNC", "Locked for Developer"),
    ("6", "PRIORITY", "Priority", "CHAR 1 (H/M/L)", "Y", "FNC", "Locked for Developer"),
    (
        "7",
        "SEVERITY",
        "Severity",
        "CHAR 1 (D/V/H/N/M)",
        "Y",
        "FNC",
        "Locked for Developer",
    ),
    (
        "8",
        "STATUS",
        "Status",
        "CHAR 20 (domain)",
        "Auto",
        "STS",
        "Always locked — via popup 0370 only",
    ),
    ("9", "DEV_ID", "Developer", "CHAR 12", "N", "EDT", "Manager only"),
    ("10", "TESTER_ID", "Tester", "CHAR 12", "N", "TST", "Manager/Tester"),
    ("11", "VERIFY_TESTER_ID", "Final Tester", "CHAR 12", "N", "EDT", "Manager only"),
    ("12", "APPROVED_BY", "Approved by", "CHAR 12", "N", "EDT", "Manager only"),
    (
        "13",
        "CREATED_AT",
        "Created Date",
        "DATS",
        "Auto",
        "—",
        "Pre-filled in Create mode",
    ),
    (
        "14",
        "CC_DESC_MINI",
        "Description (mini)",
        "Custom Control",
        "N",
        "EDT",
        "Small text editor on Bug Info tab",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws6, r, c, v, f, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws6, r, "Screen 0370 — Status Transition Popup Fields", 7)
hrow(ws6, r, ["No.", "Field", "Label", "Read/Input", "Conditional Enable", "", "Notes"])
ws6.merge_cells(f"F{r}:G{r}")
r += 1
for rd in [
    ("1", "GV_POPUP_BUG_ID", "Bug ID", "Read-only", "Always", "", ""),
    ("2", "GV_POPUP_TITLE", "Title", "Read-only", "Always", "", ""),
    ("3", "GV_POPUP_REPORTER", "Reporter (TESTER_ID)", "Read-only", "Always", "", ""),
    ("4", "GV_POPUP_CUR_STATUS", "Current Status", "Read-only", "Always", "", ""),
    (
        "5",
        "GV_POPUP_NEW_STATUS",
        "New Status",
        "Input (dropdown)",
        "Always",
        "",
        "Only allowed transitions shown",
    ),
    (
        "6",
        "GV_POPUP_DEV_ID",
        "Developer ID",
        "Input",
        "Only when ->Assigned(2)",
        "",
        "Mandatory if ->2",
    ),
    (
        "7",
        "GV_POPUP_TESTER_ID",
        "Final Tester ID",
        "Input",
        "Only when ->FinalTesting(6)",
        "",
        "Mandatory if ->6",
    ),
    (
        "8",
        "CC_TRANS_NOTE",
        "Transition Note",
        "Custom Control",
        "Mandatory for R, V",
        "",
        "Long text editor",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd[:5], 1):
        cl(ws6, r, c, v, f, BODY_FONT, LEFT)
    ws6.merge_cells(f"F{r}:G{r}")
    cl(ws6, r, 6, rd[6], f, BODY_FONT, LEFT)
    r += 1

widths(ws6, [6, 22, 22, 14, 28, 14, 18])
ws6.freeze_panes = "A4"

# ============================================================
# Sheet 7: Message Definition
# ============================================================
ws7 = wb.create_sheet("Message Definition")
ws7.merge_cells("A1:F1")
ws7["A1"].value = "Message Definition — Message Class ZBUG_MSG"
ws7["A1"].fill = HDR_FILL
ws7["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws7["A1"].alignment = CENTER
ws7.row_dimensions[1].height = 30

r = 3
hrow(ws7, r, ["Message ID", "Type", "Text (EN)", "Trigger", "Variable", "Notes"])
r += 1
msgs = [
    ("001", "S", "Bug &1 saved successfully", "Bug save OK", "&1 = BUG_ID", ""),
    ("002", "E", "Title is required", "BUG_ID missing", "—", ""),
    (
        "003",
        "E",
        "Evidence required before setting Fixed",
        "Moving to status 5 without evidence",
        "—",
        "",
    ),
    (
        "004",
        "E",
        "Transition note required for this status change",
        "Moving to R/V without TRANS_NOTE",
        "—",
        "",
    ),
    (
        "005",
        "S",
        "Status changed: &1 -> &2",
        "Status transition OK",
        "&1=old, &2=new status",
        "",
    ),
    (
        "006",
        "W",
        "No available Developer for module &1 — Bug set to Waiting",
        "Auto-assign fails",
        "&1 = SAP_MODULE",
        "",
    ),
    (
        "007",
        "S",
        "Developer &1 auto-assigned",
        "Auto-assign success",
        "&1 = DEV_ID",
        "",
    ),
    (
        "008",
        "E",
        "Cannot mark project Done: &1 open bug(s) exist",
        "Project closure validation",
        "&1 = count",
        "",
    ),
    ("009", "S", "Project &1 saved", "Project save OK", "&1 = PROJECT_ID", ""),
    (
        "010",
        "E",
        "User &1 not registered in ZBUG_USERS",
        "Invalid user",
        "&1 = user name",
        "",
    ),
    (
        "011",
        "E",
        "Access denied — role &1 cannot perform this action",
        "Permission check failed",
        "&1 = role",
        "",
    ),
    ("012", "S", "Email notification sent", "BCS email OK", "—", ""),
    (
        "013",
        "E",
        "Cannot delete: project has &1 active bug(s)",
        "Project delete validation",
        "&1 = count",
        "",
    ),
]
for rd in msgs:
    f = ALT_FILL if r % 2 == 0 else None
    type_fill = GREEN if rd[1] == "S" else (YELLOW if rd[1] == "W" else f)
    for c, v in enumerate(rd, 1):
        fill = type_fill if c == 2 else f
        cl(ws7, r, c, v, fill, BODY_FONT, CENTER if c in (1, 2) else LEFT)
    r += 1
widths(ws7, [12, 8, 45, 35, 16, 20])
ws7.freeze_panes = "A4"

# ============================================================
# Sheet 8: Processing Description
# ============================================================
ws8 = wb.create_sheet("Processing Description")
ws8.merge_cells("A1:F1")
ws8["A1"].value = "Processing Description — Auto-Assign and Role Access"
ws8["A1"].fill = HDR_FILL
ws8["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws8["A1"].alignment = CENTER
ws8.row_dimensions[1].height = 30

r = 3
r = section(
    ws8,
    r,
    "Phase A — Auto-Assign Developer (Bug Creation: New -> Assigned or Waiting)",
    6,
)
hrow(ws8, r, ["Step", "Description", "Table/Field Used", "Result", "", ""])
ws8.merge_cells(f"E{r}:F{r}")
r += 1
for rd in [
    (
        "1",
        "Trigger: Bug saved with STATUS = 1 (New)",
        "ZBUG_TRACKER.STATUS",
        "Bug save event triggered",
        "",
    ),
    (
        "2",
        "Query Developers in same PROJECT_ID",
        "ZBUG_USER_PROJEC WHERE ROLE=D AND PROJECT_ID=x",
        "Developer list obtained",
        "",
    ),
    (
        "3",
        "Filter by SAP_MODULE matching bug module",
        "JOIN ZBUG_USERS ON USER_ID",
        "Module-matching Devs",
        "",
    ),
    (
        "4",
        "Calculate workload: COUNT active bugs",
        "ZBUG_TRACKER WHERE DEV_ID=x AND STATUS IN (2,3,4,6)",
        "Workload per Dev",
        "",
    ),
    (
        "5",
        "Select Dev with lowest workload AND workload < 5",
        "Sort by workload ASC",
        "Best candidate selected",
        "",
    ),
    (
        "6a",
        "If found: set DEV_ID, STATUS=2, log history",
        "ZBUG_TRACKER.DEV_ID, ZBUG_TRACKER.STATUS",
        "STATUS=Assigned(2)",
        "",
    ),
    (
        "6b",
        "If not found: STATUS=1->W, notify Manager via email",
        "ZBUG_TRACKER.STATUS, CL_BCS",
        "STATUS=Waiting(W)",
        "",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws8, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws8, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws8, r, 3, rd[2], f, BODY_FONT, LEFT)
    ws8.merge_cells(f"D{r}:F{r}")
    cl(ws8, r, 4, rd[3], f, BODY_FONT, LEFT)
    ws8.row_dimensions[r].height = 25
    r += 1

r += 1
r = section(
    ws8, r, "Phase B — Auto-Assign Tester (Fixed -> Final Testing or Waiting)", 6
)
hrow(ws8, r, ["Step", "Description", "Table/Field Used", "Result", "", ""])
ws8.merge_cells(f"E{r}:F{r}")
r += 1
for rd in [
    (
        "1",
        "Trigger: Developer sets STATUS to 5 (Fixed) via popup 0370",
        "ZBUG_TRACKER.STATUS",
        "Fixed event triggered",
        "",
    ),
    (
        "2",
        "Query Testers in same PROJECT_ID",
        "ZBUG_USER_PROJEC WHERE ROLE=T AND PROJECT_ID=x",
        "Tester list obtained",
        "",
    ),
    (
        "3",
        "Filter by SAP_MODULE matching bug module",
        "JOIN ZBUG_USERS ON USER_ID",
        "Module-matching Testers",
        "",
    ),
    (
        "4",
        "Calculate workload: COUNT bugs in Final Testing",
        "ZBUG_TRACKER WHERE VERIFY_TESTER_ID=x AND STATUS=6",
        "Workload per Tester",
        "",
    ),
    (
        "5",
        "Select Tester with lowest workload AND workload < 5",
        "Sort by workload ASC",
        "Best candidate selected",
        "",
    ),
    (
        "6a",
        "If found: set VERIFY_TESTER_ID, STATUS=6, log history",
        "ZBUG_TRACKER.VERIFY_TESTER_ID, STATUS",
        "STATUS=FinalTesting(6)",
        "",
    ),
    (
        "6b",
        "If not found: STATUS->W, notify Manager",
        "ZBUG_TRACKER.STATUS, CL_BCS",
        "STATUS=Waiting(W)",
        "",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws8, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws8, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws8, r, 3, rd[2], f, BODY_FONT, LEFT)
    ws8.merge_cells(f"D{r}:F{r}")
    cl(ws8, r, 4, rd[3], f, BODY_FONT, LEFT)
    ws8.row_dimensions[r].height = 25
    r += 1

r += 1
r = section(ws8, r, "Role-Based Access Matrix", 6)
hrow(ws8, r, ["Action", "Manager (M)", "Developer (D)", "Tester (T)", "Notes", ""])
ws8.merge_cells(f"F{r}:F{r}")
r += 1
access = [
    ("Create Bug", "YES", "NO", "YES", ""),
    ("Delete Bug", "YES", "NO", "NO", "Soft delete IS_DEL=X"),
    (
        "Change Bug Info",
        "YES (all fields)",
        "Limited (no FNC group)",
        "Limited (no DEV group)",
        "",
    ),
    (
        "Change Status (popup)",
        "Per transition matrix",
        "Per transition matrix",
        "Per transition matrix",
        "",
    ),
    ("Upload Evidence", "YES", "YES (fix file)", "YES (report file)", ""),
    ("Download Templates", "YES", "NO", "YES", ""),
    ("View all bugs in project", "YES", "NO (own only)", "NO (own only)", ""),
    ("Create/Edit/Delete Project", "YES", "NO", "NO", ""),
    ("Add/Remove Users from Project", "YES", "NO", "NO", ""),
    ("Upload Project Excel", "YES", "NO", "NO", ""),
    ("View Project", "YES (all)", "YES (assigned)", "YES (assigned)", ""),
]
for rd in access:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws8, r, 1, rd[0], f, BODY_FONT, LEFT)
    cl(
        ws8,
        r,
        2,
        rd[1],
        GREEN if "YES" in rd[1] and "NO" not in rd[1] else f,
        BODY_FONT,
        CENTER,
    )
    cl(ws8, r, 3, rd[2], GREEN if rd[2] == "YES" else (f), BODY_FONT, CENTER)
    cl(ws8, r, 4, rd[3], GREEN if rd[3] == "YES" else (f), BODY_FONT, CENTER)
    ws8.merge_cells(f"E{r}:F{r}")
    cl(ws8, r, 5, rd[4], f, BODY_FONT, LEFT)
    ws8.row_dimensions[r].height = 22
    r += 1

widths(ws8, [32, 18, 22, 18, 32, 12])
ws8.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
