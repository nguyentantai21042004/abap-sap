import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Test_Scenario.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
ws["A1"].value = "Test Scenario"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
ws["A2"].value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Business Flow", "All cross-functional end-to-end scenarios"),
    ("Test Level", "SIT — Scenario-based integration testing"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
    ("Test Executor", "DEV-089 / DEV-061 / DEV-118"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "Scenario Legend", 8)
hrow(ws, r, ["Code", "Scenario Name", "Description", "", "", "", "", ""])
ws.merge_cells(f"D{r}:H{r}")
r += 1
for rd in [
    (
        "S1",
        "Full Happy Path",
        "One bug: New -> Assigned -> InProgress -> Fixed -> FinalTesting -> Resolved. Project marked Done.",
    ),
    (
        "S2",
        "Lifecycle with Test Failure",
        "Tester fails bug (FinalTesting -> InProgress), Dev re-fixes, Tester resolves.",
    ),
    ("S3", "Pending / Reassign", "Dev pauses bug (Pending), Manager reassigns."),
    ("S4", "Rejection Flow", "Dev rejects bug as invalid; Manager reviews."),
    (
        "S5",
        "Waiting / Manual Assign",
        "No matching Dev -> bug Waiting; Manager manually assigns.",
    ),
    ("S6", "Bug Search Flow", "Verify popup 0210 and results screen 0220."),
    ("S7", "Dashboard Accuracy", "Confirm Dashboard counts at each lifecycle step."),
    ("S8", "Auto-Assign No-Match", "No Dev/Tester match -> Waiting state correct."),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws, r, 2, rd[1], f, BOLD_FONT, LEFT)
    ws.merge_cells(f"C{r}:H{r}")
    cl(ws, r, 3, rd[2], f, BODY_FONT, LEFT)
    r += 1

widths(ws, [8, 28, 55, 12, 12, 12, 12, 12])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
ws2["A1"].value = "Change History — Test Scenario"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30
r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2,
    r,
    [
        "1",
        "1.0",
        "Initial test scenario plan — v5.0",
        "All",
        "17/04/2026",
        "DEV-089",
        "",
    ],
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: Test Scenario (Coverage Matrix)
# ============================================================
ws3 = wb.create_sheet("Test Scenario")
ws3.merge_cells("A1:J1")
ws3["A1"].value = "Test Scenario Coverage Matrix"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

r = 3
hrow(ws3, r, ["No.", "Step Name", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8"])
r += 1

steps = [
    (
        1,
        "Run ZBUG_WS -> Project Search screen (0410)",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
    ),
    (
        2,
        "Search/filter projects -> Project List (0400)",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
    ),
    (3, "Create project + add users", "Y", "", "", "", "", "", "", ""),
    (
        4,
        "Double-click project -> Bug List (0200) + Dashboard",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
        "Y",
    ),
    (5, "Create bug (Tester)", "Y", "Y", "Y", "", "", "", "", ""),
    (6, "Auto-assign Developer fires", "Y", "Y", "", "", "", "", "", "Y"),
    (7, "Assigned -> In Progress (Developer)", "Y", "Y", "Y", "", "", "", "", ""),
    (8, "In Progress -> Pending (pause)", "", "", "Y", "", "", "", "", ""),
    (9, "Pending -> Assigned again (reassign)", "", "", "Y", "", "", "", "", ""),
    (10, "Upload fix evidence (Developer)", "Y", "Y", "Y", "", "", "", "", ""),
    (11, "In Progress -> Fixed", "Y", "Y", "Y", "", "", "", "", ""),
    (12, "Auto-assign Tester fires", "Y", "Y", "", "", "", "", "", "Y"),
    (
        13,
        "Final Testing -> Resolved (Tester confirms)",
        "Y",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ),
    (14, "Final Testing -> In Progress (test fails)", "", "Y", "", "", "", "", "", ""),
    (15, "Developer rejects bug (Rejected flow)", "", "", "", "Y", "", "", "", ""),
    (16, "Manager manually assigns (Waiting flow)", "", "", "", "", "Y", "", "", ""),
    (17, "Bug Search popup (0210 -> 0220)", "", "", "", "", "", "Y", "", ""),
    (18, "Dashboard counts verified", "Y", "", "", "", "", "", "Y", ""),
    (19, "Email notification sent", "Y", "", "Y", "", "", "", "", ""),
    (
        20,
        "Project marked Done after all bugs resolved",
        "Y",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ),
]
for row in steps:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws3, r, 1, str(row[0]), f, BODY_FONT, CENTER)
    cl(ws3, r, 2, row[1], f, BODY_FONT, LEFT)
    for col, val in enumerate(row[2:], 3):
        fill = GREEN if val == "Y" else f
        cl(ws3, r, col, val, fill, BODY_FONT, CENTER)
    ws3.row_dimensions[r].height = 20
    r += 1

widths(ws3, [6, 50, 8, 8, 8, 8, 8, 8, 8, 8])
ws3.freeze_panes = "C4"

# ============================================================
# Sheet 4: Test Cases (Detailed Scenarios)
# ============================================================
ws4 = wb.create_sheet("Test Cases")
ws4.merge_cells("A1:I1")
ws4["A1"].value = "Detailed Test Cases — All Scenarios"
ws4["A1"].fill = HDR_FILL
ws4["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 30

COLS9 = [
    "NO.",
    "Step",
    "Menu Path / Screen",
    "Action",
    "Expected Result",
    "Exec",
    "Pass/Fail",
    "Notes",
    "",
]


def scen_section(ws, r, title, rows):
    r = section(ws, r, title, 9)
    hrow(ws, r, COLS9)
    r += 1
    for rd in rows:
        f = ALT_FILL if r % 2 == 0 else None
        for c, v in enumerate(rd, 1):
            cl(ws, r, c, v, f, BODY_FONT, CENTER if c in (1, 6, 7) else LEFT)
        ws.row_dimensions[r].height = 35
        r += 1
    return r + 1


r = 3

s1 = [
    (
        "1",
        "Login as Tester",
        "SAP GUI",
        "Login DEV-118",
        "SAP menu shown",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2.1",
        "Open system",
        "—",
        "Run ZBUG_WS",
        "Screen 0410 (Project Search) appears",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2.2",
        "Find project",
        "Screen 0410",
        "Execute with empty fields",
        "PRJ0000001 visible in list",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2.3",
        "Enter project",
        "Screen 0400",
        "Double-click PRJ0000001",
        "Screen 0200 Bug List + Dashboard",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "3",
        "Create bug",
        "Screen 0200",
        "Create -> TITLE=FI tax calc wrong, MODULE=FI, PRIORITY=H",
        "BUG000001 created, STATUS=Assigned, DEV_ID=DEV-061",
        "DEV-118",
        "",
        "",
        "",
    ),
    ("4", "Login as Developer", "—", "Switch to DEV-061", "—", "DEV-061", "", "", ""),
    (
        "5.1",
        "Accept bug",
        "Screen 0200",
        "Select BUG000001 -> Change -> Change Status",
        "Popup 0370 appears",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "5.2",
        "Set In Progress",
        "Screen 0370",
        "Select status=3 -> Confirm",
        "STATUS=In Progress",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "6.1",
        "Upload fix",
        "Screen 0300 Tab Evidence",
        "Upload Fix -> fix_report.xlsx",
        "Evidence row appears",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "6.2",
        "Set Fixed",
        "Screen 0300",
        "Change Status -> select 5 -> Confirm",
        "STATUS=Fixed, auto-assign Tester -> STATUS=FinalTesting(6)",
        "DEV-061",
        "",
        "",
        "",
    ),
    ("7", "Login as Tester", "—", "Switch to DEV-118", "—", "DEV-118", "", "", ""),
    (
        "8.1",
        "Open bug",
        "Screen 0200",
        "Select BUG000001 -> Change -> Change Status",
        "Popup 0370 shows status=Final Testing",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "8.2",
        "Resolve",
        "Screen 0370",
        "Select V, TRANS_NOTE=Verified all OK -> Confirm",
        "STATUS=Resolved(V)",
        "DEV-118",
        "",
        "",
        "",
    ),
    ("9", "Login as Manager", "—", "Switch to DEV-089", "—", "DEV-089", "", "", ""),
    (
        "10.1",
        "Mark project Done",
        "Screen 0400",
        "Change PRJ0000001 -> STATUS=3 (Done) -> Save",
        "PROJECT_STATUS=Done(3)",
        "DEV-089",
        "",
        "",
        "",
    ),
]
s2 = [
    (
        "1",
        "Reach Final Testing",
        "(S1 steps 1-6.2)",
        "Same as S1 up to STATUS=6",
        "STATUS=6",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "2.1",
        "Tester opens popup",
        "Screen 0300",
        "Change Status on bug in Final Testing",
        "Popup 0370, dropdown: V or 3",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2.2",
        "Fail test -> InProgress",
        "Screen 0370",
        "Select status=3, TRANS_NOTE=Field X still wrong",
        "STATUS=3, TRANS_NOTE saved, history logged",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "3.1",
        "Developer re-fixes",
        "Screen 0300",
        "Upload new fix evidence + Change Status -> 5",
        "STATUS=Fixed, auto-assign tester again",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "3.2",
        "Tester re-verifies",
        "Screen 0370",
        "Change Status -> V, TRANS_NOTE=All clear",
        "STATUS=V (Resolved)",
        "DEV-118",
        "",
        "",
        "",
    ),
]
s3 = [
    (
        "1",
        "Bug is In Progress",
        "—",
        "DEV-061 has STATUS=3",
        "—",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "2",
        "Set Pending",
        "Screen 0370",
        "Change Status -> 4",
        "STATUS=Pending",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "3",
        "Manager reassigns",
        "Screen 0370",
        "Change Status -> 2, DEV_ID=DEV-061",
        "STATUS=Assigned",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "4",
        "Dev re-accepts",
        "Screen 0370",
        "Change Status -> 3",
        "STATUS=In Progress",
        "DEV-061",
        "",
        "",
        "",
    ),
]
s4 = [
    ("1", "Bug is Assigned", "—", "STATUS=2, DEV_ID=DEV-061", "—", "—", "", "", ""),
    (
        "2",
        "Dev rejects without note",
        "Screen 0370",
        "Change Status -> R, TRANS_NOTE empty",
        "Error: note required",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "3",
        "Dev rejects with note",
        "Screen 0370",
        "Change Status -> R, TRANS_NOTE=Not a bug — by design",
        "STATUS=R, note saved",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "4",
        "History shows rejection",
        "Screen 0300 Tab History",
        "Open History tab",
        "ACTION_TYPE=ST, NEW_VALUE=R, REASON=Not a bug...",
        "DEV-089",
        "",
        "",
        "",
    ),
]
s5 = [
    (
        "1",
        "Create bug no MM Dev",
        "Screen 0200",
        "MODULE=MM, no MM Developer in project",
        "STATUS=W (Waiting), DEV_ID empty",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2",
        "Manager assigns manually",
        "Screen 0370",
        "Change Status -> 2, enter DEV_ID=DEV-061",
        "STATUS=Assigned, DEV_ID=DEV-061",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "3",
        "Continue lifecycle",
        "—",
        "Dev accepts -> In Progress -> ...",
        "Normal flow continues",
        "DEV-061",
        "",
        "",
        "",
    ),
]
s6 = [
    (
        "1",
        "Open Bug Search",
        "Screen 0200",
        "Click SEARCH button",
        "Popup 0210 appears with 6 search fields",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "2.1",
        "Search by Bug ID",
        "Popup 0210",
        "Enter BUG000001 -> Execute",
        "Screen 0220 shows 1 result",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "2.2",
        "Search by title",
        "Popup 0210",
        "TITLE=tax -> Execute",
        "All bugs with tax in title shown",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "2.3",
        "Search by status",
        "Popup 0210",
        "STATUS=3 -> Execute",
        "Only In Progress bugs",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "3",
        "View bug from results",
        "Screen 0220",
        "Select row -> Display",
        "Bug Detail (0300) opens",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "4",
        "Back from results",
        "Screen 0220",
        "Back",
        "Returns to Bug List (0200)",
        "DEV-089",
        "",
        "",
        "",
    ),
]
s7 = [
    (
        "1",
        "Initial state",
        "Screen 0200",
        "Open Bug List: 3 bugs: 1 New, 1 In Progress, 1 Resolved",
        "Total=3, New=1, InProgress=1, Resolved=1",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "2",
        "Change 1 bug to Fixed",
        "Screen 0370",
        "Status change -> 5",
        "Total=3, InProgress=0, Fixed=1, Resolved=1",
        "DEV-089",
        "",
        "",
        "",
    ),
    (
        "3",
        "After Refresh",
        "Screen 0200",
        "Click Refresh",
        "Dashboard updates to reflect new counts",
        "DEV-089",
        "",
        "",
        "",
    ),
]
s8 = [
    (
        "1",
        "No Developer for module",
        "Screen 0200",
        "Create bug, MODULE=SD, no SD Dev in project",
        "STATUS=W, DEV_ID empty, email to Manager",
        "DEV-118",
        "",
        "",
        "",
    ),
    (
        "2",
        "No Tester available",
        "—",
        "Bug reaches Fixed, no Tester available",
        "STATUS=W, VERIFY_TESTER_ID empty",
        "DEV-061",
        "",
        "",
        "",
    ),
    (
        "3",
        "Manager assigns manually",
        "Screen 0370",
        "Both cases: Manager opens popup -> assigns",
        "Status progresses correctly",
        "DEV-089",
        "",
        "",
        "",
    ),
]

r = scen_section(ws4, r, "Scenario S1 — Full Happy Path", s1)
r = scen_section(ws4, r, "Scenario S2 — Lifecycle with Test Failure", s2)
r = scen_section(ws4, r, "Scenario S3 — Pending / Reassign", s3)
r = scen_section(ws4, r, "Scenario S4 — Rejection Flow", s4)
r = scen_section(ws4, r, "Scenario S5 — Waiting / Manual Assignment", s5)
r = scen_section(ws4, r, "Scenario S6 — Bug Search Flow", s6)
r = scen_section(ws4, r, "Scenario S7 — Dashboard Accuracy", s7)
r = scen_section(ws4, r, "Scenario S8 — Auto-Assign No-Match", s8)

widths(ws4, [6, 22, 18, 40, 40, 10, 10, 14, 10])
ws4.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
