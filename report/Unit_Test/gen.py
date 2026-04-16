import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Unit_Test.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=8):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
x = ws["A1"]
x.value = "Unit Test"
x.fill = HDR_FILL
x.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
x.alignment = CENTER
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
x2 = ws["A2"]
x2.value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
x2.fill = SUB_FILL
x2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
x2.alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
info = [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Function ID", "ZBUG_WS_v5"),
    ("Test Level", "Unit Test (developer-level, pre-SIT)"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Test Executor", "DEV-089 (Manager) / DEV-061 (Developer)"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
]
for k, v in info:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "Test Accounts", 8)
hrow(ws, r, ["Account", "Role", "Role Code", "Purpose", "", "", "", ""])
ws.merge_cells(f"E{r}:H{r}")
r += 1
accounts = [
    ("DEV-089", "Manager", "M", "Tests requiring Manager role"),
    ("DEV-061", "Developer", "D", "Tests requiring Developer role"),
    ("DEV-118", "Tester", "T", "Tests requiring Tester role"),
]
for rd in accounts:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws, r, 3, rd[2], f, BODY_FONT, CENTER)
    ws.merge_cells(f"D{r}:H{r}")
    cl(ws, r, 4, rd[3], f, BODY_FONT, LEFT)
    r += 1

widths(ws, [20, 30, 12, 20, 15, 15, 15, 15])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
t = ws2["A1"]
t.value = "Change History — Unit Test"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws2.row_dimensions[1].height = 30

r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2,
    r,
    ["1", "1.0", "Initial unit test plan for v5.0", "All", "17/04/2026", "DEV-089", ""],
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: UT (Unit Test Cases)
# ============================================================
ws3 = wb.create_sheet("UT")
ws3.merge_cells("A1:H1")
t = ws3["A1"]
t.value = "Unit Test Cases — Z_BUG_WORKSPACE_MP v5.0"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws3.row_dimensions[1].height = 30

COLS = ["ID", "Test Case", "Input", "Expected Result", "Pass/Fail", "Evidence", "", ""]
r = 3


# Helper to write UT section
def ut_section(ws, r, num, name, cases):
    r = section(ws, r, f"{num}. FORM {name}", 8)
    hrow(ws, r, COLS)
    r += 1
    for rd in cases:
        f = ALT_FILL if r % 2 == 0 else None
        cl(ws, r, 1, rd[0], f, BOLD_FONT, CENTER)
        cl(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
        cl(ws, r, 3, rd[2], f, BODY_FONT, LEFT)
        cl(ws, r, 4, rd[3], f, BODY_FONT, LEFT)
        cl(ws, r, 5, "", f, BODY_FONT, CENTER)
        ws.merge_cells(f"F{r}:H{r}")
        cl(ws, r, 6, "", f, BODY_FONT, LEFT)
        ws.row_dimensions[r].height = 45
        r += 1
    return r + 1


cases1 = [
    (
        "1.1",
        "Assign Dev when suitable Dev exists",
        "BUG_ID='BUG000001', SAP_MODULE='FI', PROJECT_ID='PRJ001' — DEV-061 is FI Dev with 0 workload",
        "DEV_ID set to 'DEV-061', STATUS changes '1'->'2', ZBUG_HISTORY record inserted",
    ),
    (
        "1.2",
        "No Dev -> status Waiting",
        "SAP_MODULE='MM' — no MM Developer in project",
        "STATUS = 'W', DEV_ID empty, history log ACTION='ST'",
    ),
    (
        "1.3",
        "Dev at max workload (5) -> skip",
        "DEV-061 has 5 active bugs",
        "System skips DEV-061, tries next Dev; if none -> Waiting",
    ),
    (
        "1.4",
        "Module mismatch -> Waiting",
        "Bug module='SD', no SD Developer in project",
        "STATUS = 'W'",
    ),
]
cases2 = [
    (
        "2.1",
        "Assign Tester when available",
        "BUG_ID='BUG000001', STATUS='5' (Fixed), SAP_MODULE='FI' — DEV-118 is FI Tester with 0 workload",
        "VERIFY_TESTER_ID='DEV-118', STATUS='6' (Final Testing), history logged",
    ),
    (
        "2.2",
        "No Tester -> Waiting",
        "No Tester with matching module in project",
        "STATUS='W', VERIFY_TESTER_ID empty",
    ),
    (
        "2.3",
        "Tester at max workload (5)",
        "DEV-118 has 5 bugs in Final Testing",
        "Skip, try next; if none -> Waiting",
    ),
]
cases3 = [
    (
        "3.1",
        "Manager: New -> Assigned (valid)",
        "current=1, new=2, role=M",
        "cv_valid = TRUE",
    ),
    (
        "3.2",
        "Manager: New -> Waiting (valid)",
        "current=1, new=W, role=M",
        "cv_valid = TRUE",
    ),
    (
        "3.3",
        "Developer cannot move New -> Assigned",
        "current=1, new=2, role=D",
        "cv_valid = FALSE",
    ),
    (
        "3.4",
        "Developer: Assigned -> In Progress (valid)",
        "current=2, new=3, role=D",
        "cv_valid = TRUE",
    ),
    (
        "3.5",
        "Developer: In Progress -> Fixed (valid)",
        "current=3, new=5, role=D",
        "cv_valid = TRUE",
    ),
    (
        "3.6",
        "Tester: Final Testing -> Resolved (valid)",
        "current=6, new=V, role=T",
        "cv_valid = TRUE",
    ),
    (
        "3.7",
        "Tester: Final Testing -> In Progress (test fail)",
        "current=6, new=3, role=T",
        "cv_valid = TRUE",
    ),
    (
        "3.8",
        "Invalid direct jump New -> Resolved",
        "current=1, new=V, role=M",
        "cv_valid = FALSE",
    ),
    (
        "3.9",
        "Developer cannot reject from Resolved",
        "current=V, new=R, role=D",
        "cv_valid = FALSE",
    ),
]
cases4 = [
    (
        "4.1",
        "Create action logged",
        "ACTION_TYPE='CR', BUG_ID='BUG000001'",
        "New row in ZBUG_HISTORY with LOG_ID, CHANGED_BY=sy-uname, CHANGED_AT=sy-datum",
    ),
    (
        "4.2",
        "Status change logged",
        "ACTION_TYPE='ST', OLD_VALUE='1', NEW_VALUE='2'",
        "ZBUG_HISTORY row has correct OLD/NEW values",
    ),
    (
        "4.3",
        "LOG_ID auto-increments",
        "3 history entries for same bug",
        "LOG_IDs are sequential (MAX+1)",
    ),
    (
        "4.4",
        "Reason stored as STRING",
        "REASON='Long text with special chars'",
        "REASON saved and retrieved without truncation",
    ),
]
cases5 = [
    (
        "5.1",
        "Create new bug",
        "TITLE='Login crash', SAP_MODULE='FI', PRIORITY='H', TESTER_ID='DEV-118'",
        "Bug inserted to ZBUG_TRACKER, BUG_ID auto-generated (BUG000001), CREATED_AT=sy-datum",
    ),
    (
        "5.2",
        "BUG_ID uniqueness",
        "Create 2 bugs in sequence",
        "Second bug ID = first + 1",
    ),
    (
        "5.3",
        "Title empty -> error",
        "TITLE = ''",
        "Message E: Title is required, no INSERT",
    ),
    ("5.4", "PROJECT_ID required", "PROJECT_ID = ''", "Message E, no INSERT"),
    (
        "5.5",
        "Update existing bug",
        "BUG_ID='BUG000001', change TITLE",
        "MODIFY ZBUG_TRACKER, AENAM=sy-uname, AEDAT=sy-datum",
    ),
    (
        "5.6",
        "Soft delete",
        "Delete BUG_ID='BUG000001'",
        "IS_DEL=X set, record still in DB, not visible in list",
    ),
]
cases6 = [
    (
        "6.1",
        "Create project",
        "PROJECT_NAME='Test Project'",
        "PROJECT_ID auto-generated (PRJ0000001), PROJECT_STATUS=1",
    ),
    ("6.2", "Project name required", "PROJECT_NAME = ''", "Error message, no INSERT"),
    (
        "6.3",
        "Mark project Done — open bugs block",
        "PROJECT_STATUS='3', 2 open bugs exist",
        "Error: Cannot mark Done: 2 open bug(s)",
    ),
    (
        "6.4",
        "Mark project Done — all resolved OK",
        "All bugs STATUS='V' or '7'",
        "PROJECT_STATUS=3 saved successfully",
    ),
    ("6.5", "Soft delete project", "Delete project", "IS_DEL=X, project invisible"),
]
cases7 = [
    (
        "7.1",
        "Upload valid file",
        "BUG_ID='BUG000001', file=Bug_report.xlsx",
        "Row inserted in ZBUG_EVIDENCE, EVD_ID auto-generated, CONTENT populated",
    ),
    (
        "7.2",
        "Download uploaded file",
        "EVD_ID from 7.1",
        "File saved to local PC, size matches",
    ),
    ("7.3", "Delete evidence", "EVD_ID from 7.1", "Row deleted from ZBUG_EVIDENCE"),
]
cases8 = [
    (
        "8.1",
        "Fixed without evidence -> blocked",
        "STATUS change 3->5, ZBUG_EVIDENCE count=0 for bug",
        "Error: Evidence required before Fixed",
    ),
    (
        "8.2",
        "Fixed with evidence -> allowed",
        "ZBUG_EVIDENCE count>0",
        "Transition allowed",
    ),
    (
        "8.3",
        "Resolved without TRANS_NOTE -> blocked",
        "STATUS change 6->V, TRANS_NOTE=''",
        "Error: Transition note required",
    ),
    (
        "8.4",
        "Rejected without TRANS_NOTE -> blocked",
        "STATUS change 3->R, TRANS_NOTE=''",
        "Error: Transition note required",
    ),
]
cases9 = [
    (
        "9.1",
        "Search by Bug ID exact",
        "GV_SRCH_BUG_ID='BUG000001'",
        "GT_SEARCH_RESULTS contains exactly 1 row",
    ),
    (
        "9.2",
        "Search by title substring",
        "GV_SRCH_TITLE='login'",
        "All bugs with login in TITLE returned",
    ),
    ("9.3", "Search by status", "GV_SRCH_STATUS='3'", "Only In Progress bugs returned"),
    ("9.4", "Search by module", "GV_SRCH_MODULE='FI'", "Only FI bugs returned"),
    (
        "9.5",
        "Combined criteria",
        "GV_SRCH_MODULE='FI', GV_SRCH_STATUS='3'",
        "FI bugs that are In Progress only",
    ),
    (
        "9.6",
        "No criteria -> all project bugs",
        "All search fields empty",
        "All bugs in current project returned",
    ),
    (
        "9.7",
        "No results",
        "GV_SRCH_BUG_ID='NOTEXIST'",
        "GT_SEARCH_RESULTS empty, message shown",
    ),
]
cases10 = [
    (
        "10.1",
        "Download Bug Report template",
        "Fcode=DN_TC",
        "Bug_report.xlsx saved from SMW0 object ZBT_TMPL_01",
    ),
    (
        "10.2",
        "Download Fix template",
        "Fcode=DN_PROOF",
        "fix_report.xlsx saved from ZBT_TMPL_02",
    ),
    (
        "10.3",
        "Download Confirm template",
        "Fcode=DN_CONF",
        "confirm_report.xlsx saved from ZBT_TMPL_03",
    ),
]

r = ut_section(ws3, r, 1, "auto_assign_developer", cases1)
r = ut_section(ws3, r, 2, "auto_assign_tester", cases2)
r = ut_section(ws3, r, 3, "validate_transition", cases3)
r = ut_section(ws3, r, 4, "log_history", cases4)
r = ut_section(ws3, r, 5, "save_bug", cases5)
r = ut_section(ws3, r, 6, "save_project", cases6)
r = ut_section(ws3, r, 7, "upload_evidence", cases7)
r = ut_section(ws3, r, 8, "validate_transition — Evidence Check", cases8)
r = ut_section(ws3, r, 9, "execute_bug_search", cases9)
r = ut_section(ws3, r, 10, "download_template", cases10)

# Summary
r += 1
r = section(ws3, r, "TEST SUMMARY", 8)
hrow(ws3, r, ["Function", "Total Cases", "Pass", "Fail", "Notes", "", "", ""])
ws3.merge_cells(f"E{r}:H{r}")
r += 1
summary = [
    ("1. auto_assign_developer", "4", "", "", ""),
    ("2. auto_assign_tester", "3", "", "", ""),
    ("3. validate_transition", "9", "", "", ""),
    ("4. log_history", "4", "", "", ""),
    ("5. save_bug", "6", "", "", ""),
    ("6. save_project", "5", "", "", ""),
    ("7. upload_evidence", "3", "", "", ""),
    ("8. evidence check", "4", "", "", ""),
    ("9. execute_bug_search", "7", "", "", ""),
    ("10. download_template", "3", "", "", ""),
]
for rd in summary:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws3, r, 1, rd[0], f, BODY_FONT, LEFT)
    cl(ws3, r, 2, rd[1], f, BODY_FONT, CENTER)
    cl(ws3, r, 3, "", f, BODY_FONT, CENTER)
    cl(ws3, r, 4, "", f, BODY_FONT, CENTER)
    ws3.merge_cells(f"E{r}:H{r}")
    cl(ws3, r, 5, "", f, BODY_FONT, LEFT)
    r += 1
# Total
cl(ws3, r, 1, "Total", HDR_FILL, HDR_FONT, CENTER)
cl(ws3, r, 2, "48", HDR_FILL, HDR_FONT, CENTER)
cl(ws3, r, 3, "", HDR_FILL, HDR_FONT, CENTER)
cl(ws3, r, 4, "", HDR_FILL, HDR_FONT, CENTER)
ws3.merge_cells(f"E{r}:H{r}")
cl(ws3, r, 5, "", HDR_FILL, HDR_FONT, LEFT)

widths(ws3, [8, 38, 38, 42, 12, 12, 12, 12])
ws3.freeze_panes = "A4"

# ============================================================
# Sheet 4: Evidence
# ============================================================
ws4 = wb.create_sheet("Evidence")
ws4.merge_cells("A1:F1")
t = ws4["A1"]
t.value = "Evidence — Unit Test Screenshots"
t.fill = HDR_FILL
t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
hrow(
    ws4,
    r,
    ["Test ID", "Function", "Screenshot / Evidence File", "Result", "Date", "Tester"],
)
r += 1

evd_rows = [
    (
        "1.1",
        "auto_assign_developer",
        "SE16N ZBUG_HISTORY — new row after assign",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "1.2",
        "auto_assign_developer",
        "ZBUG_TRACKER showing STATUS=W",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "3.1",
        "validate_transition",
        "Debug watch: cv_valid=TRUE for 1->2, role=M",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "5.1",
        "save_bug",
        "SE16N ZBUG_TRACKER — new BUG000001 row",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "7.1",
        "upload_evidence",
        "ZBUG_EVIDENCE row with file data",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "8.1",
        "evidence check",
        "Error message screenshot: Evidence required",
        "Pass",
        "",
        "DEV-089",
    ),
    (
        "10.1",
        "download_template",
        "Downloaded Bug_report.xlsx on local PC",
        "Pass",
        "",
        "DEV-089",
    ),
]
for rd in evd_rows:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, LEFT)
    r += 1

# Placeholder note
ws4.merge_cells(f"A{r + 1}:F{r + 1}")
cl(
    ws4,
    r + 1,
    1,
    "Paste additional screenshots below — reference test ID in first column.",
    YELLOW,
    BODY_FONT,
    LEFT,
)

widths(ws4, [10, 28, 50, 12, 14, 14])
ws4.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
