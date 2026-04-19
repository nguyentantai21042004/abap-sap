import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "TR_Management.xlsx")

# --- Styles ---
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell(ws, r, c, val="", fill=None, font=None, align=None, border=True):
    cl = ws.cell(row=r, column=c, value=val)
    if fill:
        cl.fill = fill
    if font:
        cl.font = font
    if align:
        cl.alignment = align
    if border:
        cl.border = BORDER
    return cl


def hrow(ws, r, vals, fill=HDR_FILL, font=HDR_FONT):
    for c, v in enumerate(vals, 1):
        cell(ws, r, c, v, fill, font, CENTER)


def drow(ws, r, vals, fill=None):
    f = ALT_FILL if r % 2 == 0 else None
    if fill:
        f = fill
    for c, v in enumerate(vals, 1):
        cell(ws, r, c, v, f, BODY_FONT, LEFT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, title, subtitle=""):
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = title
    c.fill = HDR_FILL
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:K2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.fill = SUB_FILL
        c2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c2.alignment = CENTER


# ============================================================
# Sheet1 — TR Management
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Title
add_title(
    ws,
    "TR Management — Z_BUG_WORKSPACE_MP v5.0",
    "SAP System: S40 | Client: 324 | Package: ZBUGTRACK | Date: 17/04/2026",
)

# --- Section 1: Document Info ---
r = 4
ws.merge_cells(f"A{r}:K{r}")
cell(ws, r, 1, "DOCUMENT INFORMATION", SUB_FILL, SUB_FONT, CENTER)
ws.row_dimensions[r].height = 20

info = [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Transport Phase", "v5.0 Deployment (UAT -> PRD)"),
    ("SAP System", "S40 Client 324"),
    ("Package", "ZBUGTRACK"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
]
for i, (k, v) in enumerate(info):
    rr = r + 1 + i
    cell(ws, rr, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{rr}:K{rr}")
    cell(ws, rr, 2, v, None, BODY_FONT, LEFT)

r = r + 1 + len(info) + 1

# --- Section 2: TR Type Reference ---
ws.merge_cells(f"A{r}:K{r}")
cell(ws, r, 1, "TR TYPE REFERENCE", SUB_FILL, SUB_FONT, CENTER)
ws.row_dimensions[r].height = 20
r += 1
hrow(ws, r, ["Type", "Description"], SUB_FILL, SUB_FONT)
ws.merge_cells(f"B{r}:K{r}")
r += 1
drow(
    ws,
    r,
    [
        "Workbench",
        "ABAP code, screens, GUI statuses, Data Dictionary objects (SE38, SE51, SE41, SE11, SE93)",
    ],
)
ws.merge_cells(f"B{r}:K{r}")
r += 1
drow(
    ws,
    r,
    [
        "Customizing",
        "SPRO configuration (not applicable for this project — custom ABAP only)",
    ],
)
ws.merge_cells(f"B{r}:K{r}")
r += 2

# --- Section 3: Transport Request Log ---
ws.merge_cells(f"A{r}:K{r}")
cell(ws, r, 1, "TRANSPORT REQUEST LOG", SUB_FILL, SUB_FONT, CENTER)
ws.row_dimensions[r].height = 20
r += 1

hrow(
    ws,
    r,
    [
        "No.",
        "Owner",
        "Account",
        "Type",
        "TR Number",
        "Prerequisite TR",
        "Description",
        "Import By (UAT)",
        "Release Date (UAT)",
        "Import By (PRD)",
        "Release Date (PRD)",
    ],
)
ws.row_dimensions[r].height = 30
r += 1

tr_rows = [
    (
        "TR-01",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000001",
        "—",
        "DB: ZBUG_EVIDENCE table — Create table ZBUG_EVIDENCE (11 fields: BUG_ID, EVIDENCE_ID, FILE_NAME, FILE_TYPE, FILE_SIZE, UPLOADED_BY, UPLOADED_AT, CONTENT TYPE RAWSTRING, DESCRIPTION, PROJECT_ID, EVIDENCE_TYPE). Package ZBUGTRACK.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-02",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000002",
        "TR-01",
        "DB: ZBUG_TRACKER extension — Add 13 new fields: SAP_MODULE (CHAR20), BUG_TYPE, PRIORITY, SEVERITY, DEVELOPER_ID, TESTER_ID, FIXED_DATE, TESTED_DATE, TRANS_NOTE (STRING), CREATED_TIME, UPDATED_TIME + missing Data Elements/Domains.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-03",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000003",
        "TR-02",
        "DB: ZBUG_PROJECT table — Create table (16 fields: PROJECT_ID, PROJECT_NAME, DESCRIPTION, STATUS, MANAGER_ID, CREATED_BY, CREATED_AT, CREATED_TIME, UPDATED_BY, UPDATED_AT, UPDATED_TIME, PRIORITY, CATEGORY, START_DATE, END_DATE, MEMBER_COUNT).",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-04",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000004",
        "TR-03",
        "DB: ZBUG_USER_PROJEC + ZBUG_USERS extension — Create ZBUG_USER_PROJEC (10 fields). Add 4 new fields to ZBUG_USERS: EMAIL, FULL_NAME, DEPARTMENT, SKILL_LEVEL.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-05",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000005",
        "TR-04",
        "SE51: 4 new screens — Screen 0410 (Project Search, Normal), Screen 0370 (Status Transition Popup, Modal Dialog), Screen 0210 (Bug Search Input, Modal Dialog), Screen 0220 (Bug Search Results, Normal). Includes flow logic PBO/PAI.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-06",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000006",
        "TR-05",
        "SE41: GUI Statuses + Title Bars — Create STATUS_0410, STATUS_0370, STATUS_0210, STATUS_0220. Update STATUS_0200 (+SEARCH button). Create Title Bars T_0410/T_0370/T_0210/T_0220.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-07",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000007",
        "TR-06",
        "SE38: CODE v5.0 — All 6 includes: Z_BUG_WS_TOP, Z_BUG_WS_F00, Z_BUG_WS_PBO, Z_BUG_WS_PAI, Z_BUG_WS_F01, Z_BUG_WS_F02. Includes 10-state lifecycle, auto-assign, Screen 0370 popup, matrix logic, bug search, dashboard, 11 UAT bug fixes.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-08",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000008",
        "TR-07",
        "SE93: T-Code ZBUG_WS update — Change initial screen from 0400 to 0410. GUI status from STATUS_0400 to STATUS_0410.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
    (
        "TR-09",
        "DEV-089",
        "DEV-089",
        "Workbench",
        "S40K9000009",
        "TR-07",
        "SMW0: Template files — Upload 3 renamed templates: Bug_report.xlsx (ZBT_TMPL_01), fix_report.xlsx (ZBT_TMPL_02), confirm_report.xlsx (ZBT_TMPL_03). Object type: W3MIME / Binary Web Object.",
        "DEV-089",
        "",
        "DEV-089",
        "",
    ),
]

for row_data in tr_rows:
    for c, v in enumerate(row_data, 1):
        cell(ws, r, c, v, ALT_FILL if r % 2 == 0 else None, BODY_FONT, LEFT)
    ws.row_dimensions[r].height = 60
    r += 1

r += 1

# --- Section 4: Post-Import Actions ---
ws.merge_cells(f"A{r}:K{r}")
cell(ws, r, 1, "POST-IMPORT ACTIONS (Manual — Not in TR)", SUB_FILL, SUB_FONT, CENTER)
ws.row_dimensions[r].height = 20
r += 1
hrow(
    ws,
    r,
    ["Step", "Action", "Tool", "Account", "When", "", "", "", "", "", ""],
    SUB_FILL,
    SUB_FONT,
)
ws.merge_cells(f"F{r}:K{r}")
r += 1

manual = [
    (
        "M-01",
        "Activate all imported objects (activate queue)",
        "SE80 / SE38",
        "DEV-089",
        "After TR-07 import",
    ),
    (
        "M-02",
        "Run data migration: UPDATE ZBUG_TRACKER SET STATUS = V WHERE STATUS = 6",
        "SE38 / SE16",
        "DEV-089",
        "After TR-07, before UAT Round 2",
    ),
    (
        "M-03",
        "Verify ZBUG_EVIDENCE table exists and is active",
        "SE11",
        "DEV-089",
        "After TR-01 import",
    ),
    (
        "M-04",
        "Load test data: 30 mock users in ZBUG_USERS",
        "SE16",
        "DEV-089",
        "Before UAT Round 2",
    ),
    (
        "M-05",
        "Assign test users to test projects in ZBUG_USER_PROJEC",
        "SE16",
        "DEV-089",
        "Before UAT Round 2",
    ),
    (
        "M-06",
        "Smoke test: /nZBUG_WS -> Screen 0410 appears",
        "SM50",
        "DEV-089",
        "After TR-08 import",
    ),
]
for rd in manual:
    f = ALT_FILL if r % 2 == 0 else None
    cell(ws, r, 1, rd[0], f, BODY_FONT, CENTER)
    cell(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
    ws.merge_cells(f"C{r}:D{r}")
    cell(ws, r, 3, rd[2], f, BODY_FONT, CENTER)
    cell(ws, r, 5, rd[3], f, BODY_FONT, CENTER)
    cell(ws, r, 6, rd[4], f, BODY_FONT, LEFT)
    ws.merge_cells(f"F{r}:K{r}")
    r += 1

r += 1

# --- Section 5: Sign-Off Checklist ---
ws.merge_cells(f"A{r}:K{r}")
cell(ws, r, 1, "UAT SIGN-OFF CHECKLIST", SUB_FILL, SUB_FONT, CENTER)
ws.row_dimensions[r].height = 20
r += 1
hrow(
    ws,
    r,
    ["#", "Checkpoint", "Verified by", "Date", "Sign-Off", "", "", "", "", "", ""],
    SUB_FILL,
    SUB_FONT,
)
ws.merge_cells(f"E{r}:K{r}")
r += 1

checks = [
    (
        "1",
        "All TRs (TR-01 to TR-09) successfully imported to UAT without errors",
        "DEV-089",
        "",
        "",
    ),
    ("2", "Manual steps M-01 to M-06 completed", "DEV-089", "", ""),
    ("3", "Status migration executed: no STATUS=6 records remain", "DEV-089", "", ""),
    ("4", "UAT Round 2 — all 46 cases PASS", "DEV-089, DEV-061, DEV-118", "", ""),
    ("5", "11 UAT Round 1 bugs verified FIXED", "DEV-089", "", ""),
    ("6", "T-Code ZBUG_WS starts on Screen 0410 (Project Search)", "DEV-089", "", ""),
    ("7", "Auto-assign works: New -> Assigned (Developer)", "DEV-061", "", ""),
    ("8", "Auto-assign works: Fixed -> Final Testing (Tester)", "DEV-118", "", ""),
    (
        "9",
        "Status transition matrix enforced for all 3 roles",
        "DEV-089, DEV-061, DEV-118",
        "",
        "",
    ),
    ("10", "Evidence upload required before Fixed (5) transition", "DEV-118", "", ""),
]
for rd in checks:
    f = ALT_FILL if r % 2 == 0 else None
    cell(ws, r, 1, rd[0], f, BODY_FONT, CENTER)
    ws.merge_cells(f"B{r}:I{r}")
    cell(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
    cell(ws, r, 10, rd[2], f, BODY_FONT, CENTER)
    cell(ws, r, 11, rd[3], f, BODY_FONT, CENTER)
    r += 1

# PRD go-live row
cell(ws, r, 1, "PRD Go-Live Approved", GREEN, BOLD_FONT, CENTER)
ws.merge_cells(f"B{r}:I{r}")
cell(ws, r, 2, "All above checkpoints signed off", GREEN, BOLD_FONT, LEFT)
cell(ws, r, 10, "DEV-089 (Manager)", GREEN, BOLD_FONT, CENTER)
cell(ws, r, 11, "", GREEN, BOLD_FONT, CENTER)

# --- Column widths ---
set_widths(ws, [8, 12, 12, 12, 14, 14, 55, 14, 16, 14, 16])

# Freeze panes
ws.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
