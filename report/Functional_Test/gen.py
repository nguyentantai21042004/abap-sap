import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "Functional_Test.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=8):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
ws["A1"].value = "Functional Test (SIT)"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
ws["A2"].value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Function ID", "ZBUG_WS_v5"),
    ("Business Flow", "Full end-to-end bug lifecycle + project management"),
    ("Test Level", "Functional / System Integration Test (SIT)"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
    ("Test Executor", "DEV-089 (Manager), DEV-061 (Developer), DEV-118 (Tester)"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "Test Accounts", 8)
hrow(ws, r, ["Account", "Role", "Password", "Purpose", "", "", "", ""])
ws.merge_cells(f"E{r}:H{r}")
r += 1
for rd in [
    ("DEV-089", "Manager (M)", "@Anhtuoi123", "Project management, status oversight"),
    ("DEV-061", "Developer (D)", "@57Dt766", "Bug fixing, dev note, evidence"),
    ("DEV-118", "Tester (T)", "Qwer123@", "Bug creation, testing, confirmation"),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws, r, 3, rd[2], f, BODY_FONT, CENTER)
    ws.merge_cells(f"D{r}:H{r}")
    cl(ws, r, 4, rd[3], f, BODY_FONT, LEFT)
    r += 1

widths(ws, [20, 30, 14, 20, 14, 14, 14, 14])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
ws2["A1"].value = "Change History — Functional Test"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30
r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2,
    r,
    [
        "1",
        "1.0",
        "Initial SIT functional test plan — v5.0",
        "All",
        "17/04/2026",
        "DEV-089",
        "",
    ],
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: Test Cases
# ============================================================
ws3 = wb.create_sheet("Test Cases")
ws3.merge_cells("A1:H1")
ws3["A1"].value = "SIT Test Cases — Z_BUG_WORKSPACE_MP v5.0"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

TCOLS = [
    "NO.",
    "Step",
    "Test Case ID",
    "Input / Action",
    "Expected Result",
    "Exec By",
    "Pass/Fail",
    "Notes",
]

r = 3


def flow_section(ws, r, title, rows):
    r = section(ws, r, title, 8)
    hrow(ws, r, TCOLS)
    r += 1
    for rd in rows:
        f = ALT_FILL if r % 2 == 0 else None
        for c, v in enumerate(rd, 1):
            align = CENTER if c in (1, 6, 7) else LEFT
            cl(ws, r, c, v, f, BODY_FONT, align)
        ws.row_dimensions[r].height = 40
        r += 1
    return r + 1


flow1 = [
    (
        "1",
        "Create project",
        "TC-1.1",
        "PROJECT_NAME=SIT_Project_01, Manager=DEV-089",
        "PROJECT_ID auto-generated (PRJ0000001), STATUS=Opening(1)",
        "DEV-089",
        "",
        "",
    ),
    (
        "1",
        "Create project",
        "TC-1.2",
        "PROJECT_NAME empty",
        "Error: name required",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.1",
        "Add users",
        "TC-2.1",
        "USER_ID=DEV-061, ROLE=D",
        "User listed in TC_USERS table",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.2",
        "Add users",
        "TC-2.2",
        "USER_ID=DEV-118, ROLE=T",
        "User listed in TC_USERS table",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.3",
        "Add users",
        "TC-2.3",
        "Same USER_ID+PROJECT_ID again",
        "Error or graceful warning",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.4",
        "Add users",
        "TC-2.4",
        "Select DEV-061 -> Remove User",
        "DEV-061 removed from TC_USERS",
        "DEV-089",
        "",
        "",
    ),
    (
        "3.1",
        "Activate project",
        "TC-3.1",
        "PROJECT_STATUS -> 2",
        "Status updated, save success",
        "DEV-089",
        "",
        "",
    ),
    (
        "3.2",
        "Activate project",
        "TC-3.2",
        "DEV-061 tries to save project",
        "Error: access denied",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.3",
        "Complete project",
        "TC-3.3",
        "PROJECT_STATUS -> 3 (Done), 1 bug still open",
        "Error: cannot mark Done — X open bug(s)",
        "DEV-089",
        "",
        "",
    ),
    (
        "3.4",
        "Complete project",
        "TC-3.4",
        "All bugs STATUS=V -> set project Done",
        "PROJECT_STATUS=3, save OK",
        "DEV-089",
        "",
        "",
    ),
    (
        "4.1",
        "Delete project",
        "TC-4.1",
        "No bugs in project",
        "Project gone from list (IS_DEL=X)",
        "DEV-089",
        "",
        "",
    ),
]
flow2 = [
    (
        "1",
        "Create bug",
        "TC-1.1",
        "TITLE=Login fails on FI screen, MODULE=FI, PRIORITY=H, SEVERITY=V",
        "BUG_ID auto-gen, STATUS=New(1) or Assigned(2) if Dev available",
        "DEV-118",
        "",
        "",
    ),
    (
        "1",
        "Create bug",
        "TC-1.2",
        "DEV-061 is FI Dev with workload<5",
        "STATUS=Assigned(2), DEV_ID=DEV-061",
        "DEV-118",
        "",
        "",
    ),
    (
        "1",
        "Create bug",
        "TC-1.3",
        "MODULE=MM, no MM Developer in project",
        "STATUS=Waiting(W), DEV_ID empty",
        "DEV-118",
        "",
        "",
    ),
    (
        "2.1",
        "View bug",
        "TC-2.1",
        "Bug List -> Display",
        "Screen 0300 opens, all fields locked, no Save button",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.2",
        "View bug",
        "TC-2.2",
        "Click each tab",
        "Tabs Bug Info/Desc/Dev Note/Tester Note/Evidence/History all load",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.3",
        "View bug",
        "TC-2.3",
        "Open History tab after create",
        "Row with ACTION_TYPE=CR, CHANGED_BY=DEV-118",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.4",
        "View bug",
        "TC-2.4",
        "Open bug from project context",
        "PROJECT_ID matches project, field locked",
        "DEV-118",
        "",
        "",
    ),
    (
        "3.1",
        "Edit bug",
        "TC-3.1",
        "Change TITLE, MODULE, PRIORITY, SEVERITY",
        "All saved to ZBUG_TRACKER, AENAM=DEV-089",
        "DEV-089",
        "",
        "",
    ),
    (
        "3.2",
        "Edit bug",
        "TC-3.2",
        "DEV-061 changes PRIORITY",
        "Field disabled (screen group FNC)",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.3",
        "Edit bug",
        "TC-3.3",
        "DEV-118 changes PRIORITY",
        "Saved OK",
        "DEV-118",
        "",
        "",
    ),
    (
        "3.4",
        "Delete bug",
        "TC-3.4",
        "Manager deletes bug",
        "IS_DEL=X, bug gone from list",
        "DEV-089",
        "",
        "",
    ),
]
flow3 = [
    (
        "1",
        "New->Assigned",
        "TC-1.1",
        "Open popup 0370, select status=2, DEV_ID=DEV-061",
        "STATUS=2, DEV_ID=DEV-061, history ST logged",
        "DEV-089",
        "",
        "",
    ),
    (
        "1",
        "New->Assigned",
        "TC-1.2",
        "Select status=2, leave DEV_ID empty",
        "Error: developer required",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.1",
        "Assigned->InProgress",
        "TC-2.1",
        "DEV-061 opens popup, selects status=3",
        "STATUS=3, history logged",
        "DEV-061",
        "",
        "",
    ),
    (
        "2.2",
        "Assigned->InProgress",
        "TC-2.2",
        "DEV-118 tries to move Assigned->InProgress",
        "Transition blocked",
        "DEV-118",
        "",
        "",
    ),
    (
        "2.3",
        "InProgress->Rejected",
        "TC-2.3",
        "Select status=R, TRANS_NOTE empty",
        "Error: note required",
        "DEV-061",
        "",
        "",
    ),
    (
        "2.4",
        "InProgress->Rejected",
        "TC-2.4",
        "Select status=R, TRANS_NOTE=duplicate bug",
        "STATUS=R, note saved in Dev Note (Z002)",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.1",
        "InProgress->Fixed",
        "TC-3.1",
        "No evidence uploaded, select status=5",
        "Error: evidence required",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.2",
        "InProgress->Fixed",
        "TC-3.2",
        "Tab Evidence -> Upload Fix -> fix_report.xlsx",
        "File appears in evidence list",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.3",
        "InProgress->Fixed",
        "TC-3.3",
        "Select status=5",
        "STATUS=5, auto-assign Tester fires",
        "DEV-061",
        "",
        "",
    ),
    (
        "3.4",
        "InProgress->Fixed",
        "TC-3.4",
        "DEV-118 is available FI Tester",
        "STATUS=6, VERIFY_TESTER_ID=DEV-118",
        "DEV-061",
        "",
        "",
    ),
    (
        "4.1",
        "FinalTesting->Resolved",
        "TC-4.1",
        "DEV-118 selects V, TRANS_NOTE empty",
        "Error: note required",
        "DEV-118",
        "",
        "",
    ),
    (
        "4.2",
        "FinalTesting->Resolved",
        "TC-4.2",
        "Select V, TRANS_NOTE=Verified OK",
        "STATUS=V (terminal), history logged",
        "DEV-118",
        "",
        "",
    ),
    (
        "4.3",
        "FinalTesting->InProgress",
        "TC-4.3",
        "Select status=3, TRANS_NOTE=Test failed",
        "STATUS=3, log history",
        "DEV-118",
        "",
        "",
    ),
]
flow4 = [
    (
        "1",
        "Template download",
        "TC-1.1",
        "Screen 0200 -> Download Testcase",
        "Bug_report.xlsx downloaded and auto-opened",
        "DEV-089",
        "",
        "",
    ),
    (
        "1",
        "Template download",
        "TC-1.2",
        "Screen 0200 -> Download Fix",
        "fix_report.xlsx downloaded",
        "DEV-089",
        "",
        "",
    ),
    (
        "1",
        "Template download",
        "TC-1.3",
        "Screen 0200 -> Download Confirm",
        "confirm_report.xlsx downloaded",
        "DEV-118",
        "",
        "",
    ),
    (
        "2.1",
        "Upload evidence",
        "TC-2.1",
        "Tab Evidence -> Upload Report -> Bug_report.xlsx",
        "File in evidence list, ATT_REPORT updated",
        "DEV-118",
        "",
        "",
    ),
    (
        "2.2",
        "Upload evidence",
        "TC-2.2",
        "Upload Fix -> fix_report.xlsx",
        "File in list, ATT_FIX updated",
        "DEV-061",
        "",
        "",
    ),
    (
        "2.3",
        "Upload evidence",
        "TC-2.3",
        "Select row -> Download Evidence",
        "File saved to local PC",
        "DEV-089",
        "",
        "",
    ),
]
flow5 = [
    (
        "1",
        "Dashboard",
        "TC-1.1",
        "Bug List (0200) with known bug data",
        "Total = sum of all statuses; counts match SE16 data",
        "DEV-089",
        "",
        "",
    ),
    (
        "1",
        "Dashboard",
        "TC-1.2",
        "Change 1 bug status -> Refresh",
        "Counts reflect new state",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.1",
        "Bug search",
        "TC-2.1",
        "Popup 0210, Bug ID=BUG000001",
        "Screen 0220 shows exactly that bug",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.2",
        "Bug search",
        "TC-2.2",
        "Title=crash",
        "All bugs with crash in title shown",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.3",
        "Bug search",
        "TC-2.3",
        "Status=3",
        "Only In Progress bugs shown",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.4",
        "Bug search",
        "TC-2.4",
        "Bug ID=XXXXXXXX",
        "Screen 0220 empty, message shown",
        "DEV-089",
        "",
        "",
    ),
]
flow6 = [
    (
        "1",
        "Role enforcement",
        "TC-1.1",
        "Bug List -> Create button (DEV-061)",
        "Button absent or message: access denied",
        "DEV-061",
        "",
        "",
    ),
    (
        "1",
        "Role enforcement",
        "TC-1.2",
        "Select bug -> Delete (DEV-061)",
        "Error: access denied",
        "DEV-061",
        "",
        "",
    ),
    (
        "1",
        "Role enforcement",
        "TC-1.3",
        "Project List -> Create Project (DEV-118)",
        "Error: access denied",
        "DEV-118",
        "",
        "",
    ),
    (
        "1",
        "Role enforcement",
        "TC-1.4",
        "Create bug + create project (DEV-089)",
        "Both succeed",
        "DEV-089",
        "",
        "",
    ),
    (
        "2.1",
        "My Bugs filter",
        "TC-2.1",
        "DEV-061 clicks My Bugs",
        "Only bugs with DEV_ID=DEV-061 listed",
        "DEV-061",
        "",
        "",
    ),
    (
        "2.2",
        "My Bugs filter",
        "TC-2.2",
        "DEV-118 clicks My Bugs",
        "Bugs where TESTER_ID or VERIFY_TESTER_ID = DEV-118",
        "DEV-118",
        "",
        "",
    ),
]

r = flow_section(ws3, r, "Business Flow 1 — Project Lifecycle", flow1)
r = flow_section(ws3, r, "Business Flow 2 — Bug Creation and Display", flow2)
r = flow_section(ws3, r, "Business Flow 3 — Bug Lifecycle (Status Transitions)", flow3)
r = flow_section(ws3, r, "Business Flow 4 — Evidence and Templates", flow4)
r = flow_section(ws3, r, "Business Flow 5 — Dashboard and Bug Search", flow5)
r = flow_section(ws3, r, "Business Flow 6 — Access Control", flow6)

widths(ws3, [6, 18, 12, 42, 42, 10, 10, 14])
ws3.freeze_panes = "A4"

# ============================================================
# Sheet 4: Test Result
# ============================================================
ws4 = wb.create_sheet("Test Result")
ws4.merge_cells("A1:G1")
ws4["A1"].value = "Test Result Summary — SIT"
ws4["A1"].fill = HDR_FILL
ws4["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
hrow(
    ws4,
    r,
    ["Business Flow", "Total Cases", "Pass", "Fail", "Pass Rate", "Notes", "Tester"],
)
r += 1
results = [
    ("1. Project Lifecycle", "11", "", "", "", "", "DEV-089"),
    ("2. Bug Creation and Display", "11", "", "", "", "", "DEV-089/118"),
    ("3. Bug Lifecycle (Status)", "13", "", "", "", "", "DEV-089/061/118"),
    ("4. Evidence and Templates", "6", "", "", "", "", "DEV-089/061/118"),
    ("5. Dashboard and Search", "6", "", "", "", "", "DEV-089"),
    ("6. Access Control", "6", "", "", "", "", "DEV-061/118"),
]
for rd in results:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        align = CENTER if c > 1 else LEFT
        cl(ws4, r, c, v, f, BODY_FONT, align)
    r += 1
for c, v in enumerate(["Total", "53", "", "", "", "", ""], 1):
    cl(ws4, r, c, v, HDR_FILL, HDR_FONT, CENTER)

widths(ws4, [32, 14, 10, 10, 12, 24, 20])
ws4.freeze_panes = "A4"

# ============================================================
# Sheet 5: Test Data Description
# ============================================================
ws5 = wb.create_sheet("Test Data Description")
ws5.merge_cells("A1:F1")
ws5["A1"].value = "Test Data Description — SIT"
ws5["A1"].fill = HDR_FILL
ws5["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws5["A1"].alignment = CENTER
ws5.row_dimensions[1].height = 30

r = 3
hrow(ws5, r, ["Object", "ID / Value", "Purpose / Notes", "Role", "Module", "Remarks"])
r += 1
tdata = [
    ("Project", "PRJ0000001 — SIT_Project_01", "Main SIT test project", "—", "—", ""),
    ("User — Manager", "DEV-089", "All manager test cases", "M", "ABAP", ""),
    ("User — Developer", "DEV-061", "Developer test cases", "D", "FI", ""),
    ("User — Tester", "DEV-118", "Tester test cases", "T", "FI", ""),
    (
        "Bug",
        "BUG000001 — Login fails",
        "Status lifecycle tests",
        "—",
        "FI",
        "PRIORITY=H",
    ),
    (
        "Evidence file",
        "Bug_report.xlsx",
        "Downloaded from ZBT_TMPL_01",
        "—",
        "—",
        "Evidence upload tests",
    ),
    (
        "Evidence file",
        "fix_report.xlsx",
        "Downloaded from ZBT_TMPL_02",
        "—",
        "—",
        "Fix upload tests",
    ),
    (
        "Evidence file",
        "confirm_report.xlsx",
        "Downloaded from ZBT_TMPL_03",
        "—",
        "—",
        "Confirm upload tests",
    ),
]
for rd in tdata:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws5, r, c, v, f, BODY_FONT, LEFT)
    r += 1

widths(ws5, [20, 28, 42, 10, 10, 24])
ws5.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
