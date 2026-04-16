import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "UAT.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cl(ws, r, c, val="", fill=None, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    if fill:
        x.fill = fill
    if font:
        x.font = font
    if align:
        x.alignment = align
    x.border = BORDER
    return x


def hrow(ws, r, vals, fill=HDR_FILL):
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, fill, HDR_FONT, CENTER)
    ws.row_dimensions[r].height = 25


def drow(ws, r, vals):
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(vals, 1):
        cl(ws, r, c, v, f, BODY_FONT, LEFT)


def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v


def section(ws, r, title, ncols=8):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl(ws, r, 1, title, SUB_FILL, SUB_FONT, CENTER)
    ws.row_dimensions[r].height = 20
    return r + 1


wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Cover
# ============================================================
ws = wb.active
ws.title = "Cover"
ws.merge_cells("A1:H1")
ws["A1"].value = "UAT — User Acceptance Test"
ws["A1"].fill = HDR_FILL
ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
ws["A2"].value = "Z_BUG_WORKSPACE_MP — Bug Tracking System v5.0"
ws["A2"].fill = SUB_FILL
ws["A2"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 25

r = 4
for k, v in [
    ("Project Name", "Z_BUG_WORKSPACE_MP — Bug Tracking System"),
    ("Business Flow", "Full system acceptance by key users"),
    ("Test Level", "UAT — User Acceptance Test"),
    ("SAP System", "S40"),
    ("Client", "324"),
    ("T-Code", "ZBUG_WS"),
    ("Version", "1.0"),
    ("Date", "17/04/2026"),
    ("Business Key Users", "DEV-089 (Manager), DEV-061 (Developer), DEV-118 (Tester)"),
]:
    cl(ws, r, 1, k, None, BOLD_FONT, LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    cl(ws, r, 2, v, None, BODY_FONT, LEFT)
    r += 1

r += 1
r = section(ws, r, "UAT Accounts", 8)
hrow(ws, r, ["Account", "Role", "Password", "Email", "", "", "", ""])
ws.merge_cells(f"E{r}:H{r}")
r += 1
for rd in [
    ("DEV-089", "Manager (M)", "@Anhtuoi123", "TANTAISERVER2025@GMAIL.COM"),
    ("DEV-061", "Developer (D)", "@57Dt766", "HUGHNGUYEN1201@GMAIL.COM"),
    ("DEV-118", "Tester (T)", "Qwer123@", "NGUYENTANTAI.DEV@GMAIL.COM"),
]:
    f = ALT_FILL if r % 2 == 0 else None
    cl(ws, r, 1, rd[0], f, BOLD_FONT, CENTER)
    cl(ws, r, 2, rd[1], f, BODY_FONT, LEFT)
    cl(ws, r, 3, rd[2], f, BODY_FONT, CENTER)
    ws.merge_cells(f"D{r}:H{r}")
    cl(ws, r, 4, rd[3], f, BODY_FONT, LEFT)
    r += 1

widths(ws, [20, 18, 16, 32, 12, 12, 12, 12])

# ============================================================
# Sheet 2: Histories
# ============================================================
ws2 = wb.create_sheet("Histories")
ws2.merge_cells("A1:G1")
ws2["A1"].value = "Change History — UAT"
ws2["A1"].fill = HDR_FILL
ws2["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30
r = 3
hrow(
    ws2,
    r,
    [
        "No.",
        "Version",
        "Description",
        "Sheet",
        "Modified Date",
        "Modified by",
        "Remarks",
    ],
)
r += 1
drow(
    ws2, r, ["1", "1.0", "Initial UAT plan — v5.0", "All", "17/04/2026", "DEV-089", ""]
)
widths(ws2, [6, 10, 45, 10, 16, 14, 20])

# ============================================================
# Sheet 3: Test Scenario (UAT Scenarios A-H)
# ============================================================
ws3 = wb.create_sheet("Test Scenario")
ws3.merge_cells("A1:H1")
ws3["A1"].value = "UAT Test Scenarios A-H"
ws3["A1"].fill = HDR_FILL
ws3["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

UCOLS = [
    "No.",
    "Step Name",
    "Menu Path",
    "T-Code",
    "Scenario / Action",
    "Exec",
    "Pass/Fail",
    "Notes",
]


def uat_section(ws, r, title, rows):
    r = section(ws, r, title, 8)
    hrow(ws, r, UCOLS)
    r += 1
    for rd in rows:
        f = ALT_FILL if r % 2 == 0 else None
        for c, v in enumerate(rd, 1):
            cl(ws, r, c, v, f, BODY_FONT, CENTER if c in (1, 4, 6, 7) else LEFT)
        ws.row_dimensions[r].height = 35
        r += 1
    return r + 1


r = 3

uA = [
    (
        "1",
        "Start the system",
        "—",
        "ZBUG_WS",
        "Run /nZBUG_WS -> Project Search screen (0410) appears with 3 input fields",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "List all projects",
        "Screen 0410",
        "—",
        "Execute with empty filters -> Project List (0400) shows all projects",
        "DEV-089",
        "",
        "",
    ),
    (
        "3",
        "Enter Bug List",
        "Screen 0400",
        "—",
        "Double-click PRJ0000001 -> Bug List (0200) with Dashboard at top",
        "DEV-089",
        "",
        "",
    ),
    (
        "4",
        "View bug detail",
        "Screen 0200",
        "—",
        "Select bug -> Display -> Screen 0300 with 6 tabs",
        "DEV-089",
        "",
        "",
    ),
    (
        "5",
        "Navigate back",
        "Screen 0300",
        "—",
        "Back (F3) repeatedly: 0300->0200->0400->0410",
        "DEV-089",
        "",
        "",
    ),
    (
        "6",
        "Exit system",
        "Screen 0410",
        "—",
        "Back (F3) -> SAP menu",
        "DEV-089",
        "",
        "",
    ),
]
uB = [
    (
        "1",
        "Search by Project ID",
        "Screen 0410",
        "—",
        "Enter Project ID = PRJ0000001 -> Execute -> only that project shown",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "Search by Manager",
        "Screen 0410",
        "—",
        "Enter Manager = DEV-089 -> Execute -> only DEV-089 projects",
        "DEV-089",
        "",
        "",
    ),
    (
        "3",
        "Search by Status",
        "Screen 0410",
        "—",
        "Status = 2 (In Process) -> Execute -> only In Process projects",
        "DEV-089",
        "",
        "",
    ),
    (
        "4",
        "F4 lookup",
        "Screen 0410",
        "—",
        "F4 on Project ID field -> popup lists all projects -> select fills field",
        "DEV-089",
        "",
        "",
    ),
]
uC = [
    (
        "1",
        "Create project",
        "Screen 0400 -> Create Project",
        "—",
        "Fill name, manager, dates -> Save -> PROJECT_ID auto-generated",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "Add Developer",
        "Screen 0500 -> Add User",
        "—",
        "USER_ID=DEV-061, ROLE=D -> appears in table",
        "DEV-089",
        "",
        "",
    ),
    (
        "3",
        "Add Tester",
        "Screen 0500 -> Add User",
        "—",
        "USER_ID=DEV-118, ROLE=T -> appears in table",
        "DEV-089",
        "",
        "",
    ),
    (
        "4",
        "Change project status",
        "Screen 0500",
        "—",
        "STATUS: 1(Opening) -> 2(In Process) -> Save",
        "DEV-089",
        "",
        "",
    ),
    (
        "5",
        "Block Delete with bugs",
        "Screen 0400 -> Delete",
        "—",
        "Project has open bugs -> error: cannot delete",
        "DEV-089",
        "",
        "",
    ),
    (
        "6",
        "Developer cannot create project",
        "Screen 0400",
        "—",
        "DEV-061 clicks Create Project -> access denied",
        "DEV-061",
        "",
        "",
    ),
]
uD = [
    (
        "1",
        "Create bug (Tester)",
        "Screen 0200 -> Create",
        "ZBUG_WS",
        "Fill TITLE, MODULE=FI, PRIORITY=H, SEVERITY=V -> Save",
        "DEV-118",
        "",
        "",
    ),
    (
        "2",
        "PROJECT_ID pre-filled",
        "Screen 0300",
        "—",
        "PROJECT_ID = current project, field is locked",
        "DEV-118",
        "",
        "",
    ),
    (
        "3",
        "Auto-assign Dev",
        "After create",
        "—",
        "If DEV-061 is FI Dev with workload<5: STATUS=Assigned, DEV_ID=DEV-061",
        "DEV-118",
        "",
        "",
    ),
    (
        "4",
        "Edit bug (Manager)",
        "Screen 0200 -> Change",
        "—",
        "Change TITLE, PRIORITY -> Save -> updated",
        "DEV-089",
        "",
        "",
    ),
    (
        "5",
        "FNC fields locked for Dev",
        "Screen 0300 (DEV-061)",
        "—",
        "PRIORITY/SEVERITY/BUG_TYPE fields are input-disabled",
        "DEV-061",
        "",
        "",
    ),
    (
        "6",
        "Delete bug",
        "Screen 0200 -> Delete",
        "—",
        "Confirm -> bug gone from list; IS_DEL=X in DB",
        "DEV-089",
        "",
        "",
    ),
]
uE = [
    (
        "1",
        "Open status popup",
        "Screen 0300 -> Change Status",
        "—",
        "Popup 0370 shows Bug ID, Title, Reporter, Current Status + dropdown",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "New -> Assigned",
        "Popup 0370",
        "—",
        "Manager selects status=2, DEV_ID=DEV-061 -> Confirm",
        "DEV-089",
        "",
        "",
    ),
    (
        "3",
        "Assigned -> In Progress",
        "Popup 0370",
        "—",
        "DEV-061 selects status=3 -> Confirm",
        "DEV-061",
        "",
        "",
    ),
    (
        "4",
        "Upload fix evidence",
        "Tab Evidence -> Upload Fix",
        "—",
        "fix_report.xlsx uploaded",
        "DEV-061",
        "",
        "",
    ),
    (
        "5",
        "In Progress -> Fixed",
        "Popup 0370",
        "—",
        "DEV-061 selects status=5 -> auto-assign tester -> STATUS=6",
        "DEV-061",
        "",
        "",
    ),
    (
        "6",
        "Final Testing -> Resolved",
        "Popup 0370",
        "—",
        "DEV-118 selects V, TRANS_NOTE=All verified -> Confirm",
        "DEV-118",
        "",
        "",
    ),
    (
        "7",
        "Cancel popup",
        "Popup 0370",
        "—",
        "Cancel (F12) -> popup closes, status unchanged",
        "DEV-089",
        "",
        "",
    ),
    (
        "8",
        "Invalid transition blocked",
        "Popup 0370",
        "—",
        "Tester tries New->Resolved directly -> not in dropdown",
        "DEV-118",
        "",
        "",
    ),
]
uF = [
    (
        "1",
        "Download Bug Report template",
        "Screen 0200 -> Download Testcase",
        "—",
        "Bug_report.xlsx downloaded and auto-opens",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "Upload Bug Report",
        "Tab Evidence -> Upload Report",
        "—",
        "Bug_report.xlsx uploaded -> row in evidence table",
        "DEV-118",
        "",
        "",
    ),
    (
        "3",
        "Upload Fix evidence",
        "Tab Evidence -> Upload Fix",
        "—",
        "fix_report.xlsx uploaded",
        "DEV-061",
        "",
        "",
    ),
    (
        "4",
        "Upload Confirm evidence",
        "Tab Evidence -> Upload Evidence",
        "—",
        "confirm_report.xlsx uploaded",
        "DEV-118",
        "",
        "",
    ),
    (
        "5",
        "Download evidence",
        "Select row -> Download Evidence",
        "—",
        "File saved to local PC with original filename",
        "DEV-089",
        "",
        "",
    ),
    (
        "6",
        "Evidence required for Fixed",
        "Popup 0370 -> select 5 (no evidence)",
        "—",
        "Error: evidence required",
        "DEV-061",
        "",
        "",
    ),
]
uG = [
    (
        "1",
        "Dashboard visible",
        "Screen 0200",
        "—",
        "Dashboard header shows: Total, New, Assigned, InProgress, Fixed, FinalTesting, Resolved, Waiting",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "Counts correct",
        "Screen 0200",
        "—",
        "Count bugs manually -> matches Dashboard",
        "DEV-089",
        "",
        "",
    ),
    (
        "3",
        "Dashboard refreshes",
        "After status change -> Refresh",
        "—",
        "Counts update",
        "DEV-089",
        "",
        "",
    ),
    (
        "4",
        "Open Bug Search",
        "Screen 0200 -> SEARCH",
        "—",
        "Popup 0210 appears",
        "DEV-089",
        "",
        "",
    ),
    (
        "5",
        "Search by title",
        "Popup 0210 -> Title=FI -> Execute",
        "—",
        "Screen 0220 shows only FI-related bugs",
        "DEV-089",
        "",
        "",
    ),
    (
        "6",
        "Search by status",
        "Popup 0210 -> Status=3 -> Execute",
        "—",
        "Only In Progress bugs shown",
        "DEV-089",
        "",
        "",
    ),
]
uH = [
    (
        "1",
        "Send email",
        "Screen 0300 -> SENDMAIL",
        "—",
        "Message: Email sent -> check SOST",
        "DEV-089",
        "",
        "",
    ),
    (
        "2",
        "My Bugs — Developer",
        "Screen 0400 -> My Bugs",
        "—",
        "Only bugs with DEV_ID=DEV-061 listed",
        "DEV-061",
        "",
        "",
    ),
    (
        "3",
        "My Bugs — Tester",
        "Screen 0400 -> My Bugs",
        "—",
        "Bugs where TESTER_ID or VERIFY_TESTER_ID = DEV-118",
        "DEV-118",
        "",
        "",
    ),
    (
        "4",
        "Upload Project Excel",
        "Screen 0400 -> Upload",
        "—",
        "Upload valid Excel -> projects created",
        "DEV-089",
        "",
        "",
    ),
]

r = uat_section(ws3, r, "Scenario UAT-A: Access and Navigation", uA)
r = uat_section(ws3, r, "Scenario UAT-B: Project Search (Screen 0410)", uB)
r = uat_section(ws3, r, "Scenario UAT-C: Project Management", uC)
r = uat_section(ws3, r, "Scenario UAT-D: Bug Management", uD)
r = uat_section(ws3, r, "Scenario UAT-E: Status Transitions (Popup 0370)", uE)
r = uat_section(ws3, r, "Scenario UAT-F: Evidence Handling", uF)
r = uat_section(ws3, r, "Scenario UAT-G: Dashboard and Bug Search", uG)
r = uat_section(ws3, r, "Scenario UAT-H: Email, My Bugs, Templates", uH)

widths(ws3, [6, 24, 20, 10, 48, 10, 10, 18])
ws3.freeze_panes = "A4"

# ============================================================
# Sheet 4: Test Cases (Detailed UAT TC)
# ============================================================
ws4 = wb.create_sheet("Test Cases")
ws4.merge_cells("A1:H1")
ws4["A1"].value = "UAT Test Cases — Detailed"
ws4["A1"].fill = HDR_FILL
ws4["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws4["A1"].alignment = CENTER
ws4.row_dimensions[1].height = 30

r = 3
r = section(ws4, r, "UAT-TC-01: Full Happy Path", 8)
hrow(
    ws4,
    r,
    [
        "NO.",
        "Step",
        "Test Case ID",
        "Test Data",
        "Expected Result",
        "Created by",
        "Date",
        "",
    ],
)
r += 1
for rd in [
    (
        "1",
        "Create bug",
        "TC-01.1",
        "TITLE=FI tax error, MODULE=FI, PRIORITY=H",
        "BUG auto-generated, STATUS=Assigned",
        "DEV-118",
        "",
        "",
    ),
    (
        "2",
        "Accept + InProgress",
        "TC-01.2",
        "DEV-061 opens popup -> status=3",
        "STATUS=In Progress",
        "DEV-061",
        "",
        "",
    ),
    (
        "3",
        "Fix + evidence",
        "TC-01.3",
        "fix_report.xlsx -> status=5",
        "STATUS=Final Testing, Tester assigned",
        "DEV-061",
        "",
        "",
    ),
    (
        "4",
        "Resolve",
        "TC-01.4",
        "DEV-118, TRANS_NOTE=OK -> V",
        "STATUS=Resolved",
        "DEV-118",
        "",
        "",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, CENTER if c in (1, 6, 7) else LEFT)
    ws4.row_dimensions[r].height = 30
    r += 1

r += 1
r = section(ws4, r, "UAT-TC-02: Negative Cases", 8)
hrow(
    ws4,
    r,
    [
        "NO.",
        "Step",
        "Test Case ID",
        "Test Data",
        "Expected Result",
        "Created by",
        "Date",
        "",
    ],
)
r += 1
for rd in [
    (
        "1",
        "Fixed without evidence",
        "TC-02.1",
        "Status change 3->5, no evidence",
        "Error: evidence required",
        "DEV-061",
        "",
        "",
    ),
    (
        "2",
        "Resolved without note",
        "TC-02.2",
        "Status change 6->V, TRANS_NOTE empty",
        "Error: note required",
        "DEV-118",
        "",
        "",
    ),
    (
        "3",
        "Dev create bug",
        "TC-02.3",
        "DEV-061 -> Create in Bug List",
        "Access denied",
        "DEV-061",
        "",
        "",
    ),
    (
        "4",
        "Project Done with open bugs",
        "TC-02.4",
        "STATUS=3 while bugs open",
        "Error: X open bug(s)",
        "DEV-089",
        "",
        "",
    ),
]:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws4, r, c, v, f, BODY_FONT, CENTER if c in (1, 6, 7) else LEFT)
    ws4.row_dimensions[r].height = 30
    r += 1

widths(ws4, [6, 22, 14, 38, 38, 14, 14, 10])
ws4.freeze_panes = "A4"

# ============================================================
# Sheet 5: Test Result (Sign-Off)
# ============================================================
ws5 = wb.create_sheet("Test Result")
ws5.merge_cells("A1:G1")
ws5["A1"].value = "UAT Test Result — Sign-Off"
ws5["A1"].fill = HDR_FILL
ws5["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
ws5["A1"].alignment = CENTER
ws5.row_dimensions[1].height = 30

r = 3
hrow(
    ws5,
    r,
    ["Scenario", "Total Cases", "Pass", "Fail", "Pass Rate", "Sign-Off (User)", "Date"],
)
r += 1
signoffs = [
    ("A. Access and Navigation", "6", "", "", "", "DEV-089", ""),
    ("B. Project Search", "4", "", "", "", "DEV-089", ""),
    ("C. Project Management", "6", "", "", "", "DEV-089", ""),
    ("D. Bug Management", "6", "", "", "", "DEV-089/118", ""),
    ("E. Status Transitions", "8", "", "", "", "DEV-089/061/118", ""),
    ("F. Evidence", "6", "", "", "", "DEV-089/061/118", ""),
    ("G. Dashboard and Search", "6", "", "", "", "DEV-089", ""),
    ("H. Email, My Bugs", "4", "", "", "", "DEV-089/061/118", ""),
]
for rd in signoffs:
    f = ALT_FILL if r % 2 == 0 else None
    for c, v in enumerate(rd, 1):
        cl(ws5, r, c, v, f, BODY_FONT, CENTER if c > 1 else LEFT)
    r += 1
for c, v in enumerate(["Total", "46", "", "", "", "", ""], 1):
    cl(ws5, r, c, v, HDR_FILL, HDR_FONT, CENTER)
r += 2

ws5.merge_cells(f"A{r}:G{r}")
cl(
    ws5,
    r,
    1,
    "UAT PASS criteria: All 46 cases Pass (0 Fail). Any Fail must be logged in Test_And_Fix_Bug tracker, fixed, and retested before go-live approval.",
    YELLOW,
    BOLD_FONT,
    LEFT,
)
ws5.row_dimensions[r].height = 35

widths(ws5, [30, 14, 10, 10, 12, 22, 16])
ws5.freeze_panes = "A4"

wb.save(OUT)
print(f"Created: {OUT}")
